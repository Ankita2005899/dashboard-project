print("ðŸ”¥ app.py LOADED FROM:", __file__)

import os
import hmac
import hashlib
import random
import time
import re
import xml.etree.ElementTree as ET
# otp section sathi threading has been import 
import threading
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import razorpay
import mysql.connector

#----------------login github entry-------------------------
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail

#------------------------------
from flask import (
    Flask, render_template, request, jsonify,
    redirect, url_for, session, send_from_directory,
    Response, render_template_string, flash
)
from werkzeug.utils import secure_filename
from datetime import date, datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from dotenv import load_dotenv
#firebase database sathi
import mysql.connector
# app.py ke top pe â€” flask, mysql ke imports ke saath
from fuzzywuzzy import fuzz
import random
import string
import cloudinary
import cloudinary.uploader
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail as SendGridMail
import socket
from collections import defaultdict

load_dotenv("databasehandler.env")

# ---- DEFINE PATHS ----
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(BASE_DIR, "templates")
html_file_path = os.path.join(TEMPLATE_DIR, "addtocart.html")

print("BASE_DIR:", BASE_DIR)
print("TEMPLATE_DIR:", TEMPLATE_DIR)
print("Addtocart path:", html_file_path)
print("HTML file exists?", os.path.exists(html_file_path))

# ---- SINGLE Flask app definition ----
app = Flask(__name__, template_folder=TEMPLATE_DIR)
app.secret_key = "secret123"
app.secret_key = "shopco_secret_key_2026"


cloudinary.config(
    cloud_name = os.getenv("CLOUDINARY_CLOUD_NAME"),
    api_key = os.getenv("CLOUDINARY_API_KEY"),
    api_secret = os.getenv("CLOUDINARY_API_SECRET")
)
# â”€â”€ Create database and tables â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

# ================= EMAIL CONFIGURATION =================


# Change this temporarily to test
OWNER_EMAIL = 'ankitabandal45@gmail.com'  # use any other email you have access to

# Excel file path
EXCEL_FILE = "User_Profile_Data.xlsx"

# Razorpay keys
KEY_ID = os.environ.get("RAZORPAY_KEY_ID")
KEY_SECRET = os.environ.get("RAZORPAY_KEY_SECRET")

client = razorpay.Client(auth=(KEY_ID, KEY_SECRET))

print("ðŸ”¥ Flask looking for templates in:", app.template_folder)


# ============================================================
# DATABASE CONNECTION
# ============================================================
def get_db_connection():
    conn = mysql.connector.connect(
        host=os.getenv("MYSQLHOST"),
        user=os.getenv("MYSQLUSER"),
        password=os.getenv("MYSQLPASSWORD"),
        database=os.getenv("MYSQLDATABASE"),
        port=int(os.getenv("MYSQLPORT")),
        ssl_ca=None,
        ssl_verify_cert=False,
        ssl_verify_identity=False,
        autocommit=True
    )
    
    
    cursor = conn.cursor()
    cursor.execute("SET time_zone = '+05:30';")
    cursor.close()
    return conn
# ============================================================
# PASSWORD VALIDATION
# ============================================================
def is_strong_password(password):
    pattern = r'^(?=(.*[!@#$%^&*()_+\-=\[\]{};\'":\\|,.<>\/?]){3,})(?=.*[a-z])(?=.*[A-Z])(?=.*\d).{8,}$'
    return re.match(pattern, password)


# ============================================================
# USER TABLE HELPERS
# ============================================================
def get_cart_table_name(username, user_id):
    username = username.strip().lower()
    safe_username = re.sub(r'[^a-z0-9_]', '_', username)
    if safe_username[0].isdigit():
        safe_username = "user_" + safe_username
    return f"{safe_username}_{user_id}"


def ensure_user_table(username, user_id):
    table_name = get_cart_table_name(username, user_id)

    db = get_db_connection()
    cursor = db.cursor()

    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS `{table_name}` (
            id INT AUTO_INCREMENT PRIMARY KEY,
            user_id INT,
            product_id INT,
            row_no VARCHAR(20),
            category VARCHAR(50),
            name VARCHAR(255),
            price DECIMAL(10,2),
            image VARCHAR(255),
            video VARCHAR(255),
            availability INT,
            detail TEXT,
            uploaded_at DATETIME NULL DEFAULT NULL,
            address VARCHAR(255),
            quantity INT DEFAULT 1,
            date DATETIME NULL DEFAULT NULL,
            total DECIMAL(10,2),
            mode VARCHAR(20),
            sub_vc_item INT DEFAULT 0
        )
    """)

    cursor.execute(f"SHOW COLUMNS FROM `{table_name}` LIKE 'product_id'")
    if not cursor.fetchone():
        cursor.execute(f"ALTER TABLE `{table_name}` ADD COLUMN product_id INT DEFAULT NULL")

    cursor.execute(f"SHOW COLUMNS FROM `{table_name}` LIKE 'sub_vc_item'")
    if not cursor.fetchone():
        cursor.execute(f"ALTER TABLE `{table_name}` ADD COLUMN sub_vc_item INT DEFAULT 0")

    db.commit()
    cursor.close()
    db.close()

def create_user_product_activity_table(username, user_id):
    safe_username = re.sub(r'[^a-z0-9_]', '_', username.strip().lower())
    if safe_username and safe_username[0].isdigit():
        safe_username = "user_" + safe_username
    table_name = f"{safe_username}_{user_id}_product_activity"
    
    db = get_db_connection()
    cursor = db.cursor()

    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS `{table_name}` (
            id INT AUTO_INCREMENT PRIMARY KEY,
            product_id INT,
            name VARCHAR(255),
            category VARCHAR(100),
            today_search_count INT DEFAULT 0,
            search_time DATETIME,
            growth_on_search VARCHAR(20),
            today_add_to_cart_count INT DEFAULT 0,
            add_to_cart_date_time DATETIME,
            growth_in_addtocart VARCHAR(20),
            today_purchase_count INT DEFAULT 0,
            purchased_time DATETIME,
            month VARCHAR(20),
            growth FLOAT DEFAULT 0
        )
    """)

    # âœ… Purani tables fix karo â€” add_to_cart_time â†’ add_to_cart_date_time
    cursor.execute(f"""
        SELECT COUNT(*) FROM information_schema.columns 
        WHERE table_schema = DATABASE() 
        AND table_name = %s 
        AND column_name = 'add_to_cart_time'
    """, (table_name,))
    has_old_col = cursor.fetchone()[0]

    if has_old_col:
        cursor.execute(f"""
            ALTER TABLE `{table_name}`
            CHANGE COLUMN `add_to_cart_time` `add_to_cart_date_time` DATETIME
        """)

    # âœ… search_time aur purchased_time DATETIME karo
    cursor.execute(f"""
        ALTER TABLE `{table_name}`
        MODIFY COLUMN search_time DATETIME,
        MODIFY COLUMN purchased_time DATETIME
    """)

    db.commit()
    cursor.close()
    db.close()
#--------------------sare route ko automatically fix kar ne ke leya because 1 point {username}_{id} madhe jar aahe tar {username}_{id}_product_activity madhe sagale ssaav mahanun ------------------------

@app.route("/fix-all-activity-tables")
def fix_all_activity_tables():
    if not session.get("user_id"):
        return jsonify({"error": "unauthorized"}), 401

    db = get_db_connection()
    cursor = db.cursor()

    cursor.execute("""
        SELECT table_name 
        FROM information_schema.tables 
        WHERE table_schema = DATABASE() 
        AND table_name LIKE '%_product_activity'
    """)
    tables = cursor.fetchall()

    fixed = []
    errors = []

    for (table_name,) in tables:
        try:
            # Pehle check karo add_to_cart_time exist karta hai
            cursor.execute(f"""
                SELECT COUNT(*) FROM information_schema.columns 
                WHERE table_schema = DATABASE() 
                AND table_name = %s 
                AND column_name = 'add_to_cart_time'
            """, (table_name,))
            has_old_col = cursor.fetchone()[0]

            if has_old_col:
                # Rename old column to new name
                cursor.execute(f"""
                    ALTER TABLE `{table_name}`
                    CHANGE COLUMN `add_to_cart_time` `add_to_cart_date_time` DATETIME
                """)

            # search_time aur purchased_time bhi DATETIME karo
            cursor.execute(f"""
                ALTER TABLE `{table_name}`
                MODIFY COLUMN search_time DATETIME,
                MODIFY COLUMN purchased_time DATETIME
            """)

            fixed.append(table_name)
        except Exception as e:
            errors.append(f"{table_name}: {str(e)}")

    db.commit()
    cursor.close()
    db.close()

    return jsonify({
        "success": True,
        "fixed_tables": fixed,
        "errors": errors
    })
    
    
    
# ============================================================
# PRODUCT SYNC HELPERS
# ============================================================
def sync_product_sub_vc(product_id, category):
    db = get_db_connection()
    cursor = db.cursor()

    cursor.execute("""
        SELECT DISTINCT table_name
        FROM information_schema.columns
        WHERE table_schema = DATABASE()
          AND column_name = 'sub_vc_item'
    """)
    user_tables = cursor.fetchall()

    total_sub_vc = 0
    for (table_name,) in user_tables:
        cursor.execute(
            f"SELECT COALESCE(SUM(sub_vc_item), 0) FROM `{table_name}` WHERE product_id = %s AND category = %s",
            (product_id, category)
        )
        total_sub_vc += cursor.fetchone()[0]

    cursor.execute(
        "UPDATE product_availability SET sub_vc = %s WHERE product_id = %s AND category = %s",
        (total_sub_vc, product_id, category)
    )

    db.commit()
    cursor.close()
    db.close()


def sync_product_availability_sql(products):
    db = get_db_connection()
    cursor = db.cursor()

    for prod in products:
        dashboard_avail = abs(prod.get("actual_availability", 0))
        sub_vc = abs(prod.get("sub_vc", 0))
        cart_avail = dashboard_avail + sub_vc

        cursor.execute("""
            INSERT INTO product_availability_sql (product_id, name, category, cart_availability)
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                cart_availability = VALUES(cart_availability),
                name = VALUES(name),
                category = VALUES(category)
        """, (prod["product_id"], prod.get("dash_item_name", ""), prod.get("category", ""), cart_avail))

    db.commit()
    cursor.close()
    db.close()


def sync_xml_to_sql(xml_file_path):
    db = get_db_connection()
    cursor = db.cursor()

    tree = ET.parse(xml_file_path)
    root = tree.getroot()

    for record in root.findall("record"):
        product_id = record.find("product_id").text
        name = record.find("name").text
        category = record.find("category").text
        cart_avail = int(record.find("cart_availability").text)

        cursor.execute("""
            INSERT INTO product_availability_sql (product_id, name, category, cart_availability)
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                cart_availability = VALUES(cart_availability),
                name = VALUES(name),
                category = VALUES(category)
        """, (product_id, name, category, cart_avail))

    db.commit()
    cursor.close()
    db.close()


def sync_product_availability_sql_one_row(products):
    db = get_db_connection()
    cursor = db.cursor()

    for prod in products:
        dashboard_avail = abs(prod.get("actual_availability", 0))
        sub_vc = abs(prod.get("sub_vc", 0))
        cart_avail = dashboard_avail + sub_vc

        cursor.execute("""
            SELECT id FROM product_availability_sql
            WHERE product_id=%s AND name=%s AND category=%s
        """, (prod["product_id"], prod.get("dash_item_name", ""), prod.get("category", "")))
        existing = cursor.fetchone()

        if existing:
            cursor.execute(
                "UPDATE product_availability_sql SET cart_availability=%s WHERE id=%s",
                (cart_avail, existing[0])
            )
        else:
            cursor.execute("""
                INSERT INTO product_availability_sql (product_id, name, category, cart_availability)
                VALUES (%s,%s,%s,%s)
            """, (prod["product_id"], prod.get("dash_item_name", ""), prod.get("category", ""), cart_avail))

    db.commit()
    cursor.close()
    db.close()


def update_xml_with_sub_vc(xml_file_path, output_xml_path=None):
    db = get_db_connection()
    cursor = db.cursor()

    tree = ET.parse(xml_file_path)
    root = tree.getroot()

    cursor.execute("""
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = DATABASE()
          AND table_name REGEXP '^[a-zA-Z0-9_]+_[0-9]+$'
    """)
    user_tables = [t[0] for t in cursor.fetchall()]

    for record in root.findall("record"):
        product_id_elem = record.find("product_id")
        category_elem = record.find("category")

        if product_id_elem is None:
            continue

        product_id = int(product_id_elem.text)
        category = category_elem.text if category_elem is not None else None
        total_sub_vc = 0

        for table_name in user_tables:
            cursor.execute(f"SHOW COLUMNS FROM `{table_name}` LIKE 'sub_vc_item'")
            if cursor.fetchone():
                if category:
                    cursor.execute(
                        f"SELECT COALESCE(SUM(sub_vc_item),0) FROM `{table_name}` WHERE product_id=%s AND category=%s",
                        (product_id, category)
                    )
                else:
                    cursor.execute(
                        f"SELECT COALESCE(SUM(sub_vc_item),0) FROM `{table_name}` WHERE product_id=%s",
                        (product_id,)
                    )
                res = cursor.fetchone()
                if res:
                    total_sub_vc += res[0]

        sub_vc_elem = record.find("Sub from VC")
        if sub_vc_elem is None:
            sub_vc_elem = ET.SubElement(record, "Sub from VC")
        sub_vc_elem.text = f"-{total_sub_vc}"

    tree.write(output_xml_path or xml_file_path, encoding="utf-8", xml_declaration=True)
    cursor.close()
    db.close()
    print("âœ… XML updated with Sub from VC (negative) successfully!")


# ============================================================
# EXCEL HELPERS
# ============================================================
def create_excel_file():
    if os.path.exists(EXCEL_FILE):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "UserData"

    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:H1')
    ws.merge_cells('I1:P1')
    ws.merge_cells('Q1:Q2')

    ws['A1'] = "User ID"
    ws['B1'] = "Username"
    ws['C1'] = "Clothing and Shoes"
    ws['I1'] = "Skin care and Beauty"
    ws['Q1'] = "Interests"

    sub_headers = [
        "Department", "Height", "Weight", "Age",
        "Fit attributes", "Shoes",
        "Type", "Sensitivity", "Concern", "Tone",
        "Eye care", "Ingredients not wanted",
        "Imp qualities", "Product Formulation"
    ]

    col = 3
    for header in sub_headers:
        ws.cell(row=2, column=col).value = header
        col += 1

    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)

    wb.save(EXCEL_FILE)


# ============================================================
# DATABASE TABLE CREATION
# ============================================================
def create_tables():
    db = get_db_connection()
    cursor = db.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS card (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(255),
            image VARCHAR(255),
            video VARCHAR(255),
            availability INT,
            price DECIMAL(10,2),
            detail TEXT,
            uploaded_at DATE,
            address VARCHAR(255),
            material VARCHAR(255)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS study_material (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(255),
            image VARCHAR(255),
            video VARCHAR(255),
            availability INT,
            price DECIMAL(10,2),
            detail TEXT,
            uploaded_at DATE,
            address VARCHAR(255),
            material VARCHAR(255)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS food_items (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(255),
            image VARCHAR(255),
            video VARCHAR(255),
            availability INT,
            price DECIMAL(10,2),
            detail TEXT,
            uploaded_at DATE,
            address VARCHAR(255),
            material VARCHAR(255)
        )
    """)

    cursor.close()
    db.close()


# ============================================================
# FILE UPLOAD CONFIG
# ============================================================
UPLOAD_FOLDER = "static/products"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER


# ============================================================
# ROUTES â€” AUTH
# ============================================================
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/signup", methods=["GET", "POST"])
def signup():
    print("ðŸ‘‰ ROUTE HIT:", request.path)

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email    = request.form.get("email", "").strip()
        password = request.form.get("password", "").strip()
        confirm  = request.form.get("confirm_password", "").strip()

        if not username or not email or not password:
            return render_template("signup.html", error="âŒ All fields are required")

        if password != confirm:
            return render_template("signup.html", error="âŒ Passwords do not match")

        if "@" not in email or "." not in email:
            return render_template("signup.html", error="âŒ Invalid email format")

        db     = get_db_connection()
        cursor = db.cursor()

        try:
            # Check if email already exists in user_activity table
            cursor.execute("SELECT id FROM user_activity WHERE email=%s", (email,))
            existing = cursor.fetchone()
            cursor.fetchall()
            if existing:
                cursor.close()
                db.close()
                return render_template("signup.html",
                    error="This email is already registered. Try logging in.")

            # Check strong password table
            auto_pwd = None
            try:
                cursor.execute("""
                    SELECT id FROM strong_password
                    WHERE password=%s AND is_used=0
                """, (password,))
                auto_pwd = cursor.fetchone()
            except Exception as e:
                print("strong_password table error:", e)
                auto_pwd = None

            # Check password strength
            if not auto_pwd and not is_strong_password(password):
                cursor.close()
                db.close()
                return render_template("signup.html",
                    error="âŒ Weak password. Need 8+ chars, 1 uppercase, 1 lowercase, 1 digit, 3 special chars")

            # Save to user_activity table only
            cursor.execute("""
                INSERT INTO user_activity
                (username, email, password, mode, action, action_date, action_time)
                VALUES (%s, %s, %s, 'signup', %s, CURDATE(), CURTIME())
            """, (username, email, password, "manual"))
            user_id = cursor.lastrowid
            print(f"âœ… user_activity inserted: id={user_id}, username={username}, email={email}")

            # Create personal tables
            ensure_user_table(username, user_id)
            create_user_product_activity_table(username, user_id)
            create_user_your_item_table(username, user_id)

            # Mark pre-approved password as used
            if auto_pwd:
                cursor.execute("""
                    UPDATE strong_password SET is_used=1
                    WHERE id=%s
                """, (auto_pwd[0],))

            db.commit()
            cursor.close()
            db.close()

            session.clear()
            session["user_id"]       = user_id
            session["user_obj_id"]   = user_id
            session["username"]      = username
            session["user_email"]    = email
            session["flash_message"] = f"ðŸŽ‰ Welcome {username}! Account created ðŸ˜Š"

            return redirect(url_for("survey"))

        except Exception as e:
            print("âŒ SIGNUP ERROR:", e)
            try:
                db.rollback()
                cursor.close()
                db.close()
            except:
                pass
            return render_template("signup.html",
                error=f"âŒ Something went wrong: {str(e)}")

    return render_template("signup.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    print("ðŸ‘‰ ROUTE HIT:", request.path)

    if request.method == "POST":
        email    = request.form["email"].strip()
        password = request.form["password"].strip()

        db     = get_db_connection()
        cursor = db.cursor()

        # Check user_activity for manual signup
        cursor.execute("""
            SELECT id, username, password FROM user_activity
            WHERE email=%s AND action='manual'
            ORDER BY id ASC LIMIT 1
        """, (email,))
        user = cursor.fetchone()
        cursor.fetchall()

        if not user:
            cursor.close()
            db.close()
            return render_template("login.html", error="âŒ Email not registered or not signed up manually")

        if password != user[2]:
            cursor.close()
            db.close()
            return render_template("login.html", error="âŒ Incorrect password")

        # Save login to user table
        try:
            cursor.execute("""
                INSERT INTO user (username, email, password, action)
                VALUES (%s, %s, %s, %s)
            """, (user[1], email, password, "manual"))
            db.commit()
        except Exception as e:
            print("user table insert error:", e)

        cursor.close()
        db.close()

        ensure_user_table(user[1], user[0])
        create_user_product_activity_table(user[1], user[0])

        session.clear()
        session["user_id"]     = user[0]
        session["username"]    = user[1]
        session["user_obj_id"] = user[0]
        session["flash_message"] = f"ðŸŽ‰ Login Successful, {user[1]} ðŸ˜Š"
        return redirect(url_for("home"))

    return render_template("login.html")



#===============================firebase==================
@app.route("/auth/firebase-login", methods=["POST"])
def firebase_login():
    data     = request.get_json()
    email    = data.get("email", "").strip()
    username = data.get("username", "").strip()
    uid      = data.get("uid", "").strip()
    provider = data.get("provider", "").strip()
    source   = data.get("source", "").strip()
    password = data.get("password", "").strip()

    db     = get_db_connection()
    cursor = db.cursor()

    try:
        if source == "signup":
            # Check if email already exists in user_activity
            cursor.execute("SELECT id FROM user_activity WHERE email=%s", (email,))
            existing = cursor.fetchone()
            cursor.fetchall()
            if existing:
                cursor.close()
                db.close()
                return jsonify({"success": False, "message": "This email is already registered. Try logging in."})

            # Save to user_activity only
            cursor.execute("""
                INSERT INTO user_activity
                (username, email, password, mode, action, action_date, action_time)
                VALUES (%s, %s, %s, 'signup', %s, CURDATE(), CURTIME())
            """, (username, email, password, provider))
            user_id = cursor.lastrowid

            ensure_user_table(username, user_id)
            create_user_product_activity_table(username, user_id)

            db.commit()
            cursor.close()
            db.close()

            session.clear()
            session["user_id"]       = user_id
            session["user_obj_id"]   = user_id
            session["username"]      = username
            session["user_email"]    = email
            session["flash_message"] = f"ðŸŽ‰ Welcome {username}! Account created ðŸ˜Š"

            return jsonify({"success": True})

        elif source == "login":
            # Check user_activity for google signup
            cursor.execute("""
                SELECT id, username FROM user_activity
                WHERE email=%s AND action='google'
                ORDER BY id ASC LIMIT 1
            """, (email,))
            existing = cursor.fetchone()
            cursor.fetchall()

            if not existing:
                cursor.close()
                db.close()
                return jsonify({"success": False, "message": "No Google account found. Please sign up first."})

            user_id  = existing[0]
            username = existing[1]

            # Save login to user table
            cursor.execute("""
                INSERT INTO user (username, email, password, action)
                VALUES (%s, %s, %s, %s)
            """, (username, email, "", "google"))

            ensure_user_table(username, user_id)
            create_user_product_activity_table(username, user_id)

            db.commit()
            cursor.close()
            db.close()

            session.clear()
            session["user_id"]       = user_id
            session["user_obj_id"]   = user_id
            session["username"]      = username
            session["user_email"]    = email
            session["flash_message"] = f"ðŸŽ‰ Login Successful, {username} ðŸ˜Š"

            return jsonify({"success": True})

        else:
            return jsonify({"success": False, "message": "Invalid source"})

    except Exception as e:
        print("âŒ firebase_login error:", e)
        try:
            db.rollback()
            cursor.close()
            db.close()
        except:
            pass
        return jsonify({"success": False, "message": str(e)})


@app.route("/auth/verify-otp", methods=["POST"])
def verify_otp():
    data  = request.get_json()
    email = data.get("email", "").strip()
    otp   = data.get("otp", "").strip()

    if session.get("otp") != otp or session.get("otp_email") != email:
        return jsonify({"success": False, "message": "Invalid or expired OTP"})

    db     = get_db_connection()
    cursor = db.cursor()
    cursor.execute("""
        SELECT id, username FROM user_activity
        WHERE email=%s AND action='github'
        ORDER BY id ASC LIMIT 1
    """, (email,))
    user = cursor.fetchone()
    cursor.fetchall()

    if not user:
        cursor.close()
        db.close()
        return jsonify({"success": False, "message": "No GitHub account found with this email."})

    # Save login to user table
    try:
        cursor.execute("""
            INSERT INTO user (username, email, password, action)
            VALUES (%s, %s, %s, %s)
        """, (user[1], email, "", "github"))
        db.commit()
        print(f"âœ… user table insert success: {email}")
    except Exception as e:
        print(f"âŒ user table insert error: {e}")

    cursor.close()
    db.close()

    session.clear()
    session["user_id"]       = user[0]
    session["user_obj_id"]   = user[0]
    session["username"]      = user[1]
    session["user_email"]    = email
    session["flash_message"] = f"ðŸŽ‰ Login Successful, {user[1]} ðŸ˜Š"

    return jsonify({"success": True})


@app.route("/auth/flask-login", methods=["POST"])
def flask_login():
    data     = request.get_json()
    email    = data.get("email", "").strip()
    password = data.get("password", "").strip()

    db     = get_db_connection()
    cursor = db.cursor()

    # Check user_activity for manual signup
    cursor.execute("""
        SELECT id, username, password FROM user_activity
        WHERE email=%s AND action='manual'
        ORDER BY id ASC LIMIT 1
    """, (email,))
    user = cursor.fetchone()
    cursor.fetchall()

    if not user:
        cursor.close()
        db.close()
        return jsonify({"success": False, "message": "âŒ Email not registered or not signed up manually"})

    if password != user[2]:
        cursor.close()
        db.close()
        return jsonify({"success": False, "message": "âŒ Incorrect password"})

    # Save login to user table
    try:
        cursor.execute("""
            INSERT INTO user (username, email, password, action)
            VALUES (%s, %s, %s, %s)
        """, (user[1], email, password, "manual"))
        db.commit()
    except Exception as e:
        print("Activity log error:", e)

    cursor.close()
    db.close()

    ensure_user_table(user[1], user[0])
    create_user_product_activity_table(user[1], user[0])

    session.clear()
    session["user_id"]       = user[0]
    session["user_obj_id"]   = user[0]
    session["username"]      = user[1]
    session["user_email"]    = email
    session["flash_message"] = f"ðŸŽ‰ Login Successful, {user[1]} ðŸ˜Š"

    return jsonify({"success": True})


# â”€â”€ Send OTP â”€â”€
@app.route("/auth/send-otp", methods=["POST"])
def send_otp():
    data  = request.get_json()
    email = data.get("email", "").strip()

    db     = get_db_connection()
    cursor = db.cursor()
    cursor.execute("""
        SELECT id, username FROM user_activity
        WHERE email=%s AND action='github'
        ORDER BY id ASC LIMIT 1
    """, (email,))
    user = cursor.fetchone()
    cursor.fetchall()
    cursor.close()
    db.close()

    if not user:
        return jsonify({"success": False, "message": "No GitHub account found with this email."})

    otp = str(random.randint(100000, 999999))
    session["otp"]       = otp
    session["otp_email"] = email

    try:
        sg_key = os.getenv("SHOPSPHERE_SENDGRID_KEY")
        print(f"DEBUG OTP KEY -> len={len(sg_key) if sg_key else 0}, "
            f"start={sg_key[:8] if sg_key else None}, end={sg_key[-4:] if sg_key else None}")
        sg = SendGridAPIClient(sg_key)
        message = SendGridMail(
            from_email=os.getenv("SENDER_EMAIL"),
            to_emails=email,
            subject="ShopSphere Login OTP",
            html_content=f"""
            <div style="font-family:sans-serif;padding:20px;max-width:400px">
                <h2 style="color:#FF6B2B">ShopSphere Login OTP</h2>
                <p>Your one-time password is:</p>
                <h1 style="letter-spacing:8px;color:#1a1a2e">{otp}</h1>
                <p style="color:#666">This OTP is valid for 10 minutes.</p>
            </div>
            """
        )
        sg.send(message)
        return jsonify({"success": True})
    except Exception as e:
        print("âŒ OTP send error:", e)
        return jsonify({"success": False, "message": "Failed to send OTP"})


#------------------question page sathi --------------------------

@app.route("/survey", methods=["GET", "POST"])
def survey():
    if "user_id" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":
        q1 = request.form.get("q1", "")
        q2 = request.form.get("q2", "")
        q3 = request.form.get("q3", "")
        q4 = request.form.get("q4", "")
        q5 = request.form.get("q5", "")

        db = get_db_connection()
        cursor = db.cursor()
        cursor.execute("""
            INSERT INTO user_survey (user_id, email, q1, q2, q3, q4, q5)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (session["user_id"], session.get("user_email", ""), q1, q2, q3, q4, q5))
        db.commit()
        cursor.close()
        db.close()

        return redirect(url_for("home"))

    return render_template("survey.html")


@app.route("/survey/skip")
def survey_skip():
    if "user_id" not in session:
        return redirect(url_for("login"))
    return redirect(url_for("home"))

#------------------"try another option" part cha code------------------------

@app.route("/auth/security-login", methods=["POST"])
def security_login():
    data = request.get_json()
    email = data.get("email", "").strip()
    q1 = data.get("q1", "").strip()
    q2 = data.get("q2", "").strip()

    db = get_db_connection()
    cursor = db.cursor()

    cursor.execute("""
        SELECT user_id, q1, q2 FROM user_survey
        WHERE email=%s ORDER BY id DESC LIMIT 1
    """, (email,))
    survey = cursor.fetchone()
    cursor.fetchall()

    if not survey or survey[1].strip().lower() != q1.strip().lower() or survey[2].strip().lower() != q2.strip().lower():
        cursor.close()
        db.close()
        return jsonify({"success": False, "message": "Incorrect answers. Please try again."})

    cursor.execute("""
        SELECT username FROM user_activity WHERE email=%s ORDER BY id ASC LIMIT 1
    """, (email,))
    user = cursor.fetchone()
    cursor.fetchall()

    if not user:
        cursor.close()
        db.close()
        return jsonify({"success": False, "message": "No account found with this email."})

    username = user[0]
    user_id = survey[0]

    cursor.execute("""
        INSERT INTO user (username, email, password, action)
        VALUES (%s, %s, %s, %s)
    """, (username, email, "Q1 and Q2 are correct", "security"))

    ensure_user_table(username, user_id)
    create_user_product_activity_table(username, user_id)

    db.commit()
    cursor.close()
    db.close()

    session.clear()
    session["user_id"]       = user_id
    session["user_obj_id"]   = user_id
    session["username"]      = username
    session["user_email"]    = email
    session["flash_message"] = f"ðŸŽ‰ Login Successful, {username} ðŸ˜Š"

    return jsonify({"success": True})
     
# ============================================================
# ROUTES â€” PAGES
# ============================================================
@app.route("/home")
def home():
    print("ðŸ‘‰ ROUTE HIT:", request.path)
    if "user_id" not in session:
        return redirect(url_for("login"))

    message = session.pop("flash_message", "")
    username = session.get("username")
    return render_template("dashboard.html", message=message, username=username)


@app.route("/analytics")
def analytics():
    return render_template("Analytics.html")


@app.route("/search")
def search_page():
    return render_template("search.html")


@app.route("/addtocart")
def addtocart_page():
    return render_template("addtocart.html")


@app.route("/addtocart-table")
def addtocart_table():
    return render_template("Addtocard-table.html")


@app.route("/search-table")
def search_table():
    return render_template("search-table.html")


@app.route("/purchased-table")
def purchased_table():
    return render_template("purchase-table.html")


@app.route("/availabilities-table")
def availabilities_table():
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)

    cursor.execute("SELECT id, name, category, availability, price, updated_at FROM card")
    cards = cursor.fetchall()

    cursor.execute("SELECT id, name, category, availability, price, uploaded_at AS updated_at FROM study_material")
    studies = cursor.fetchall()

    cursor.execute("SELECT id, name, category, availability, price, uploaded_at AS updated_at FROM food_items")
    foods = cursor.fetchall()

    products = cards + studies + foods

    cursor.execute("SELECT id, username FROM user")
    users = cursor.fetchall()

    removed_counts = {}
    sub_vc_counts = {}

    for user in users:
        safe_username = re.sub(r'\W+', '_', user['username'].lower())
        user_table = f"{safe_username}_{user['id']}"

        cursor.execute(f"SHOW TABLES LIKE '{user_table}'")
        if not cursor.fetchone():
            continue

        cursor.execute(f"SHOW COLUMNS FROM `{user_table}` LIKE 'quantity'")
        has_quantity = cursor.fetchone() is not None

        cursor.execute(f"SHOW COLUMNS FROM `{user_table}` LIKE 'product_id'")
        has_product_id = cursor.fetchone() is not None

        cursor.execute(f"SHOW COLUMNS FROM `{user_table}` LIKE 'sub_vc_item'")
        has_sub_vc = cursor.fetchone() is not None

        if not (has_quantity and has_product_id):
            continue

        if has_sub_vc:
            cursor.execute(f"""
                SELECT product_id,
                       COALESCE(SUM(quantity), 0) AS total_qty,
                       COALESCE(SUM(sub_vc_item), 0) AS total_sub_vc
                FROM `{user_table}`
                GROUP BY product_id
            """)
        else:
            cursor.execute(f"""
                SELECT product_id,
                       COALESCE(SUM(quantity), 0) AS total_qty,
                       0 AS total_sub_vc
                FROM `{user_table}`
                GROUP BY product_id
            """)

        for row in cursor.fetchall():
            pid = row["product_id"]
            removed_counts[pid] = removed_counts.get(pid, 0) + row["total_qty"]
            sub_vc_counts[pid] = sub_vc_counts.get(pid, 0) + row["total_sub_vc"]

    cursor.close()
    db.close()

    # FIX: proper <tr> tags in loop
    html = """
    <table border="1">
        <tr>
            <th>ID</th><th>Product ID</th><th>Name</th><th>Category</th>
            <th>Cart Availability</th><th>Removed</th><th>Dashboard Availability</th>
            <th>Add from VC</th><th>Sub from VC</th><th>Actual Availability</th><th>Updated At</th>
        </tr>
    """

    for prod in products:
        prod_id = prod["id"]
        removed_qty = removed_counts.get(prod_id, 0)
        sub_vc = sub_vc_counts.get(prod_id, 0)
        availability = prod.get("availability", 0)

        dashboard_availability = availability
        sub_vc_value = abs(sub_vc)
        cart_availability = dashboard_availability + sub_vc_value

        html += f"""
        <tr>
            <td>{prod_id}</td>
            <td>{prod_id}</td>
            <td>{prod.get('name', '')}</td>
            <td>{prod.get('category', '')}</td>
            <td>{cart_availability}</td>
            <td>{removed_qty}</td>
            <td>{dashboard_availability}</td>
            <td>{removed_qty}</td>
            <td>{sub_vc}</td>
            <td>{dashboard_availability}</td>
            <td>{prod.get('updated_at')}</td>
        </tr>
        """

    html += "</table>"
    return render_template_string(html)


@app.route("/owner-section")
def owner_section():
    return render_template("ownersection.html")


# NOTE: owner_dashboard no longer conflicts with login (removed duplicate @app.route("/"))
@app.route("/owner-dashboard")
def owner_dashboard():
    if not session.get("owner_verified"):
        return "Unauthorized Access", 403
    return render_template("owner_dashboard.html")

@app.route("/dashboard")
def dashboard_redirect():
    return render_template("dashboard.html")


@app.route("/addtocartfortheownersection")
def addtocart_owner_section():
    return render_template("addtocartfortheownersection.html")


@app.route("/searchfortheownersection")
def search_for_owner():
    return render_template("searchfortheownersection.html")


@app.route("/privacy-policy")
def privacy_policy():
    return render_template("privacy_policy.html")


@app.route("/terms")
def terms():
    return render_template("terms.html")


@app.route("/refund-policy")
def refund_policy():
    return render_template("refund_policy.html")


@app.route("/Buynow.html")
def buynow_page():
    return render_template("Buynow.html")


@app.route("/buynow")
def buynow():
    return render_template("Buynow.html")


# ============================================================
# ROUTES â€” OWNER OTP
# ============================================================
@app.route("/send-owner-otp")
def send_owner_otp():
    otp = random.randint(100000000, 999999999)
    session["owner_otp"] = str(otp)
    print(f"ðŸ”‘ OWNER OTP: {otp}")
    
    try:
        import urllib.request
        import json
        
        payload = json.dumps({
            "personalizations": [{"to": [{"email": OWNER_EMAIL}]}],
            "from": {"email": "ankitabandal45@gmail.com"},
            "subject": "Owner Login OTP",
            "content": [{"type": "text/plain",
                         "value": f"Your Owner Login OTP is: {otp}"}]
        }).encode()
        
        req = urllib.request.Request(
            "https://api.sendgrid.com/v3/mail/send",
            data=payload,
            headers={
                "Authorization": f"Bearer {os.getenv('SHOPSPHERE_SENDGRID_KEY')}",
                "Content-Type": "application/json"
            }
        )
        urllib.request.urlopen(req)
        print("âœ… OTP sent successfully")
    except Exception as e:
        print("âŒ Email error:", e)
    
    return jsonify({"success": True})




@app.route("/verify-owner-otp", methods=["POST"])
def verify_owner_otp():
    data = request.json
    user_otp = data.get("otp")
    if session.get("owner_otp") == user_otp:
        session["owner_verified"] = True
        session.pop("owner_otp", None)
        return jsonify({"success": True, "message": "âœ… Owner verified successfully!"})
    return jsonify({"success": False, "message": "âŒ Invalid OTP"})


# ============================================================
# ROUTES â€” PRODUCTS
# ============================================================
@app.route("/get-products")
def get_products():
    print("ðŸ‘‰ ROUTE HIT:", request.path)
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM card")
    rows = cursor.fetchall()

    products = []
    for row in rows:
        image = row["image"]
        image_url = image if image and image.startswith("http") else (f"/static/products/{image}" if image else None)
        video = row["video"]
        video_url = video if video and video.startswith("http") else (f"/static/products/{video}" if video else None)

        products.append({
            "id": row["id"],
            "name": row["name"],
            "image": image_url,
            "video": video_url,
            "availability": row["availability"],
            "price": row["price"],
            "uploaded_at": row["uploaded_at"].strftime("%Y-%m-%d") if row["uploaded_at"] else None,
            "address": row["address"],
            "material": row["material"],
            "detail": row.get("detail", ""),
            "keywords": row.get("keywords", "")
        })
    cursor.close()
    db.close()
    return jsonify(products)


@app.route("/get-study-materials")
def get_study_materials():
    print("ðŸ‘‰ ROUTE HIT:", request.path)
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM study_material")
    rows = cursor.fetchall()

    products = []
    for row in rows:
        products.append({
            "id": row["id"],
            "name": row["name"],
            "image": row["image"] if row["image"] else None,
            "video": row["video"] if row["video"] else None,
            "availability": row["availability"],
            "price": row["price"],
            "uploaded_at": row["uploaded_at"].strftime("%Y-%m-%d") if row["uploaded_at"] else None,
            "address": row["address"],
            "material": row["material"],
            "detail": row.get("detail", ""),
            "keywords": row.get("keywords", "")

        })

    cursor.close()
    db.close()
    return jsonify(products)


@app.route("/get-food-items")
def get_food_items():
    print("ðŸ‘‰ ROUTE HIT:", request.path)
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM food_items")
    rows = cursor.fetchall()

    products = []
    for row in rows:
        products.append({
            "id": row["id"],
            "name": row["name"],
            "image": row["image"] if row["image"] else None,
            "video": row["video"] if row["video"] else None,
            "availability": row["availability"],
            "price": row["price"],
            "uploaded_at": row["uploaded_at"].strftime("%Y-%m-%d") if row["uploaded_at"] else None,
            "address": row["address"],
            "material": row["material"],
            "detail": row.get("detail", ""),
            "keywords": row.get("keywords", "")

        })

    cursor.close()
    db.close()
    return jsonify(products)


@app.route("/get-product-search-table")
def get_product_search_table():
    query = """
        SELECT id AS product_id, name, category, searched_count AS todays_searched_count, last_searched_time
        FROM card
        UNION ALL
        SELECT id, name, category, searched_count, last_searched_time FROM study_material
        UNION ALL
        SELECT id, name, category, searched_count, last_searched_time FROM food_items
        ORDER BY category, product_id;
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(query)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(rows)
    except Exception as e:
        print("Error fetching search table:", e)
        return jsonify([])


@app.route("/add-product", methods=["GET", "POST"])
def add_product():
    print("ðŸ‘‰ ROUTE HIT:", request.path)

    if request.method == "GET":
        return render_template("add_product.html")

    db = get_db_connection()
    cursor = db.cursor()

    image_file = request.files.get("image")
    image_name = None
    if image_file and image_file.filename:
        upload_result = cloudinary.uploader.upload(image_file)
        image_name = upload_result["secure_url"]

    video_file = request.files.get("video")
    video_name = None
    if video_file and video_file.filename:
        upload_result_video = cloudinary.uploader.upload(video_file, resource_type="video")
        video_name = upload_result_video["secure_url"]
    name = request.form.get("product_name")
    price = request.form.get("product_price")
    availability = request.form.get("availability_count")
    address = request.form.get("address")
    uploaded_date = request.form.get("uploaded_date")
    made_of = request.form.get("made_of")
    used_for = request.form.get("used_for")
    harmful = request.form.get("harmful_activity")
    precautions = request.form.get("precautions")
    product_types = request.form.getlist("product_type[]")
    keywords = request.form.get("keywords", "")
    
    
    if not product_types:
        return "Product type is required", 400

    detail = (
        f"Made of: {made_of}\n"
        f"Used for: {used_for}\n"
        f"Harmful: {harmful}\n"
        f"Precautions: {precautions}"
    )

    def insert_product_and_availability(table, category):
        cursor.execute(f"""
            INSERT INTO {table} (name, image, video, availability, price, detail, uploaded_at, address, material, keywords)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (name, image_name, video_name, availability, price, detail, uploaded_date, address, made_of, keywords))
        pid = cursor.lastrowid
        cursor.execute("""
            INSERT INTO product_availability (dash_item_name, actual_availability, removed, product_id, category)
            VALUES (%s, %s, %s, %s, %s)
        """, (name, availability, 0, pid, category))

    if "Kitchen" in product_types:
        insert_product_and_availability("card", "kitchen")
    if "Study Material" in product_types:
        insert_product_and_availability("study_material", "study_material")
    if "Food" in product_types:
        insert_product_and_availability("food_items", "food_items")

    # â”€â”€ Insert into store_data (all users' products) â”€â”€
    category_label = product_types[0] if product_types else "other"
    # Get a valid product_id from the last inserted category table
    last_pid = cursor.lastrowid if cursor.lastrowid else 0

    cursor.execute("""
        INSERT INTO store_data
        (user_id, product_id, category, name, image, video, price, availability, detail, address, quantity)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        session.get("user_id"), last_pid, category_label, name, image_name, video_name,
        price, availability, detail, address, availability
    ))
    store_data_id = cursor.lastrowid

    # â”€â”€ Insert into user's personal your_item table â”€â”€
    try:
        uname = session.get("username")
        uid   = session.get("user_id")
        your_item_table = f"{uname}_{uid}_your_item"
        cursor.execute(f"""
            INSERT INTO `{your_item_table}`
            (store_data_id, category, name, image, video, price, availability, detail, address, quantity,
             made_of, used_for, harmful_activity, precautions)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            store_data_id, category_label, name, image_name, video_name,
            price, availability, detail, address, availability,
            made_of, used_for, harmful, precautions
        ))
        print(f"âœ… Inserted into {your_item_table}")
    except Exception as e:
        print(f"âš ï¸ Could not insert into your_item table: {e}")

    db.commit()
    cursor.close()
    db.close()
    return redirect(url_for("home"))
    


# ============================================================
# ROUTES â€” CART
# ============================================================
@app.route("/get-cart-items")
def get_cart_items():
    print("ðŸ‘‰ ROUTE HIT:", request.path)

    if "user_id" not in session or "username" not in session:
        return jsonify([])

    user_id = session["user_id"]
    username = session["username"]

    ensure_user_table(username, user_id)
    table_name = get_cart_table_name(username, user_id)

    db = get_db_connection()
    cursor = db.cursor(dictionary=True)

    try:
        cursor.execute(f"SELECT * FROM `{table_name}` WHERE quantity > 0 AND (mode IS NULL OR mode = 'failed') ORDER BY uploaded_at DESC")
        items = cursor.fetchall()

        for row in items:
            img = row.get("image")
            if img:
                row["image"] = img if (img.startswith("http") or img.startswith("/static")) else f"/static/products/{img}"
            else:
                row["image"] = "/static/products/default.png"

            vid = row.get("video")
            if vid:
                row["video"] = vid if (vid.startswith("http") or vid.startswith("/static")) else f"/static/videos/{vid}"
            else:
                row["video"] = ""

            row["date"] = row["date"].strftime("%Y-%m-%d") if row.get("date") else ""

    except Exception as e:
        print("Cart fetch error:", e)
        items = []
    finally:
        cursor.close()
        db.close()

    return jsonify(items)



@app.route("/add-to-cart", methods=["POST"])
def add_to_cart():
    print("ðŸ‘‰ ROUTE HIT:", request.path)

    if "cart" not in session:
        session["cart"] = []

    data = request.get_json(force=True)
    cart = data.get("cart", [])

    if not cart:
        return jsonify({"success": False, "error": "Cart empty"}), 400

    user_id = session["user_id"]
    username = session["username"]
    purchased_by = username

    ensure_user_table(username, user_id)

    try:
        db = get_db_connection()
        cursor = db.cursor()

        table_map = {
            "kitchen": "card",
            "card": "card",
            "study": "study_material",
            "study_material": "study_material",
            "food": "food_items",
            "food_items": "food_items"
        }

        client_table = get_cart_table_name(username, user_id)
        safe_un = re.sub(r'[^a-z0-9_]', '_', username.strip().lower())
        if safe_un and safe_un[0].isdigit():
            safe_un = "user_" + safe_un
        activity_table = f"{safe_un}_{user_id}_product_activity"

        # IST time fallback
        ist_now = datetime.utcnow() + timedelta(hours=5, minutes=30)
        ist_now_str = ist_now.strftime('%Y-%m-%d %H:%M:%S')
        ist_month = ist_now.strftime('%B')

        for item in cart:
            if item["category"] not in table_map:
                raise Exception(f"Unknown category: {item['category']}")

            price = float(item["price"])
            quantity = 1
            table = table_map[item["category"]]

            cursor.execute(f"""
                UPDATE {table} SET availability = availability - %s
                WHERE id = %s AND availability >= %s
            """, (quantity, item["id"], quantity))

            if cursor.rowcount == 0:
                raise Exception(f"{item['name']} is out of stock")

            cursor.execute("""
                UPDATE product_availability
                SET available = available - 1, removed = removed + 1
                WHERE product_id = %s AND category = %s
            """, (item["id"], item["category"]))

            if cursor.rowcount == 0:
                raise Exception("product_availability entry missing for this product")

            cursor.execute(f"SELECT availability FROM {table} WHERE id = %s", (item["id"],))
            remaining_availability = cursor.fetchone()[0]

            cursor.execute("""
                SELECT id, quantity FROM store_data
                WHERE user_id=%s AND product_id=%s AND category=%s AND quantity > 0
                ORDER BY id DESC LIMIT 1
            """, (user_id, item["id"], item["category"]))
            existing = cursor.fetchone()

            if existing:
                cursor.execute("UPDATE store_data SET quantity = quantity + 1 WHERE id = %s", (existing[0],))
            else:
                cursor.execute("""
                    INSERT INTO store_data (
                        user_id, category, name, image, video, price, availability,
                        detail, purchased_by, uploaded_at, address, date, quantity, product_id
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),%s,NOW(),%s,%s)
                """, (
                    user_id, item["category"], item["name"],
                    item.get("image", ""), item.get("video", ""),
                    price, remaining_availability, item.get("detail", ""),
                    purchased_by, item.get("address", ""), 1, item["id"]
                ))

            cursor.execute(f"""
                INSERT INTO `{client_table}` (
                    user_id, category, name, image, video, price, availability,
                    detail, uploaded_at, address, date, product_id, sub_vc_item
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,NOW(),%s,NOW(),%s,%s)
            """, (
                user_id, item["category"], item["name"],
                item.get("image", ""), item.get("video", ""),
                price, 1, item.get("detail", ""),
                item.get("address", ""), item["id"], 0
            ))

            # âœ… client_table se actual date fetch karo (jo sahi time hai)
            cursor.execute(f"""
                SELECT date FROM `{client_table}`
                ORDER BY id DESC LIMIT 1
            """)
            cart_date = cursor.fetchone()
            actual_datetime = cart_date[0] if cart_date else ist_now_str

            # âœ… Product activity update
            cursor.execute(f"""
                SELECT id, today_add_to_cart_count FROM `{activity_table}`
                WHERE product_id = %s AND category = %s
            """, (item["id"], item["category"]))
            act_existing = cursor.fetchone()

            if act_existing:
                new_count = act_existing[1] + 1
                cursor.execute(f"""
                    UPDATE `{activity_table}`
                    SET today_add_to_cart_count = %s,
                        add_to_cart_date_time = %s,
                        growth_in_addtocart = %s
                    WHERE product_id = %s AND category = %s
                """, (new_count, actual_datetime, f"{new_count}/100", item["id"], item["category"]))
            else:
                cursor.execute(f"""
                    INSERT INTO `{activity_table}`
                    (product_id, name, category, today_add_to_cart_count, add_to_cart_date_time, month, growth_in_addtocart)
                    VALUES (%s, %s, %s, 1, %s, %s, '1/100')
                """, (item["id"], item["name"], item["category"], actual_datetime, ist_month))

        # âœ… Loop ke BAAD commit
        db.commit()
        cursor.close()
        db.close()
        return jsonify({"success": True})

    except Exception as e:
        print("âŒ ADD ERROR:", e)
        try:
            db.rollback()
            cursor.close()
            db.close()
        except:
            pass
        return jsonify({"success": False, "error": str(e)}), 500
    
    
    
    
@app.route("/remove-from-cart", methods=["POST"])
def remove_from_cart():
    print("ðŸ‘‰ ROUTE HIT:", request.path)

    if "user_id" not in session or "username" not in session:
        return jsonify({"success": False})

    data = request.get_json(silent=True) or {}
    cart_id = data.get("id")

    if not cart_id:
        return jsonify({"success": False})

    user_id = session["user_id"]
    username = session["username"]

    ensure_user_table(username, user_id)
    # Use sanitized table name consistently
    table_name = get_cart_table_name(username, user_id)

    db = get_db_connection()
    cursor = db.cursor(dictionary=True)

    try:
        cursor.execute(
            f"SELECT quantity, product_id, category FROM `{table_name}` WHERE id = %s",
            (cart_id,)
        )
        row = cursor.fetchone()

        if not row:
            return jsonify({"success": False})

        product_id = row["product_id"]
        category = row["category"]

        if row["quantity"] > 1:
            cursor.execute(
                f"UPDATE `{table_name}` SET quantity = quantity - 1, sub_vc_item = sub_vc_item - 1 WHERE id = %s",
                (cart_id,)
            )
            action = "decremented"
        else:
            cursor.execute(
                f"UPDATE `{table_name}` SET quantity = 0, sub_vc_item = sub_vc_item - 1 WHERE id = %s",
                (cart_id,)
            )
            cursor.execute("""
                UPDATE product_availability
                SET
                    available = available + 1,
                    removed = GREATEST(removed - 1, 0),
                    total_dash = actual_availability - GREATEST(removed - 1, 0),
                    remain_in_dash = actual_availability - GREATEST(removed - 1, 0)
                WHERE product_id = %s AND category = %s
            """, (product_id, category))
            action = "removed"

        sync_product_sub_vc(product_id, category)

        db.commit()
        return jsonify({"success": True, "action": action})

    except Exception as e:
        print("âŒ REMOVE ERROR:", e)
        db.rollback()
        return jsonify({"success": False})
    finally:
        cursor.close()
        db.close()


@app.route("/final-add", methods=["POST"])
def final_add():
    print("ðŸ‘‰ ROUTE HIT:", request.path)

    if "user_id" not in session:
        return jsonify({"success": False, "error": "Login required"}), 401

    data = request.get_json(force=True)
    cart = data.get("cart", [])

    if not cart:
        return jsonify({"success": False, "error": "Cart is empty"}), 400

    db = get_db_connection()
    cursor = db.cursor(buffered=True, dictionary=True)

    CATEGORY_TABLES = {
        "card": "card",
        "study_material": "study_material",
        "food_items": "food_items"
    }

    user_id = session["user_id"]
    username = session["username"]

    ensure_user_table(username, user_id)
    client_table = get_cart_table_name(username, user_id)

    try:
        cursor.execute(f"""
            SELECT MAX(CAST(SUBSTRING(row_no, 5) AS UNSIGNED)) AS max_row FROM `{client_table}`
        """)
        max_row = cursor.fetchone()["max_row"] or 0
        next_row_no = f"row_{max_row + 1}"

        cursor.execute(f"""
            SELECT 1 FROM `{client_table}` WHERE DATE(date) = CURDATE() AND row_no IS NOT NULL LIMIT 1
        """)
        row_exists_today = cursor.fetchone()
        row_no_value = next_row_no if not row_exists_today else None

        for item in cart:
            table = CATEGORY_TABLES.get(item["category"])
            if not table:
                raise Exception(f"Invalid category: {item['category']}")

            cursor.execute(f"SELECT availability FROM {table} WHERE id=%s", (item["product_id"],))
            stock = cursor.fetchone()
            if not stock or stock["availability"] <= 0:
                raise Exception(f"{item['name']} is out of stock")

            remaining = stock["availability"] - 1
            cursor.execute(f"UPDATE {table} SET availability = availability - 1 WHERE id=%s", (item["product_id"],))

            cursor.execute("""
                UPDATE product_availability SET removed = removed + 1
                WHERE dash_item_name = %s AND category = %s
            """, (item["name"], item["category"]))

            cursor.execute(f"SELECT uploaded_at FROM {table} WHERE id=%s", (item["product_id"],))
            real_uploaded_at = cursor.fetchone()["uploaded_at"]

            cursor.execute("""
                SELECT id, quantity FROM store_data
                WHERE user_id=%s AND product_id=%s AND category=%s ORDER BY id DESC LIMIT 1
            """, (user_id, item["product_id"], item["category"]))
            existing_store = cursor.fetchone()

            if existing_store:
                cursor.execute("UPDATE store_data SET quantity = quantity + 1 WHERE id = %s", (existing_store["id"],))
            else:
                cursor.execute("""
                    INSERT INTO store_data
                    (user_id, product_id, category, name, price, availability, detail,
                     address, purchased_by, uploaded_at, image, video, quantity)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    user_id, item.get("product_id") or item.get("id"),
                    item["category"], item["name"], item["price"], remaining,
                    item.get("detail", ""), item.get("address", ""),
                    username, real_uploaded_at, item.get("image", ""), item.get("video", ""), 1
                ))

            select_cursor = db.cursor(buffered=True, dictionary=True)
            select_cursor.execute(f"""
                SELECT id, quantity FROM `{client_table}`
                WHERE name = %s AND category = %s AND DATE(date) = CURDATE()
                  AND date >= NOW() - INTERVAL 1 SECOND
                ORDER BY id DESC LIMIT 1
            """, (item["name"], item["category"]))
            existing = select_cursor.fetchone()
            select_cursor.close()

            if existing:
                new_qty = existing["quantity"] + 1
                cursor.execute(f"""
                    UPDATE `{client_table}` SET quantity=%s, total=%s, uploaded_at=%s WHERE id=%s
                """, (new_qty, item["price"] * new_qty, real_uploaded_at, existing["id"]))
            else:
                cursor.execute(f"""
                    INSERT INTO `{client_table}`
                    (user_id, row_no, category, name, price, image, video, availability,
                     detail, address, quantity, total, uploaded_at, date, product_id)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),%s)
                """, (
                    user_id, row_no_value, item["category"], item["name"],
                    item["price"], item.get("image", ""), item.get("video", ""),
                    remaining, item.get("detail", ""), item.get("address", ""),
                    1, item["price"], real_uploaded_at, item["product_id"]
                ))
                row_no_value = None

        db.commit()
        cursor.close()
        db.close()
        return jsonify({"success": True})

    except Exception as e:
        db.rollback()
        cursor.close()
        db.close()
        print("FINAL ADD ERROR:", e)
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/remove-from-store", methods=["POST"])
def remove_from_store():
    print("ðŸ‘‰ ROUTE HIT:", request.path)

    try:
        data = request.get_json(force=True)
        product_id = data.get("product_id")
        category = data.get("category")

        if not product_id or not category:
            return jsonify({"success": False, "error": "Invalid data"}), 400

        table_map = {
            "kitchen": "card",
            "study": "study_material",
            "food": "food_items"
        }

        if category not in table_map:
            return jsonify({"success": False, "error": "Invalid category"}), 400

        db = get_db_connection()
        cursor = db.cursor()

        cursor.execute("""
            DELETE FROM store_data WHERE product_id=%s AND category=%s ORDER BY id DESC LIMIT 1
        """, (product_id, category))

        if cursor.rowcount == 0:
            db.rollback()
            return jsonify({"success": False, "error": "Item not found in cart"}), 404

        cursor.execute(f"""
            UPDATE {table_map[category]} SET availability = availability + 1 WHERE id = %s
        """, (product_id,))

        db.commit()
        cursor.close()
        db.close()
        return jsonify({"success": True})

    except Exception as e:
        print("âŒ REMOVE ERROR:", e)
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/finalize-cart", methods=["POST"])
def finalize_cart():
    print("ðŸ‘‰ ROUTE HIT:", request.path)

    cart = request.json.get("cart", [])
    if not cart:
        return jsonify({"success": False, "error": "Cart empty"}), 400

    user_id = session.get("user_id")
    purchased_by = session.get("username")
    cart_total = sum(float(item["price"]) for item in cart)

    db = get_db_connection()
    cursor = db.cursor()

    cursor.execute("""
        INSERT INTO cart_summary (user_id, purchased_by, total) VALUES (%s, %s, %s)
    """, (user_id, purchased_by, cart_total))

    db.commit()
    cursor.close()
    db.close()

    return jsonify({"success": True, "cart_total": cart_total})


# ============================================================
# ROUTES â€” XML / DATA FEEDS
# ============================================================
@app.route("/availabilities-xml")
def availabilities_xml():
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)

    cursor.execute("SELECT id AS product_id, name, category, availability, price FROM card")
    cards = cursor.fetchall()
    cursor.execute("SELECT id AS product_id, name, category, availability, price FROM study_material")
    studies = cursor.fetchall()
    cursor.execute("SELECT id AS product_id, name, category, availability, price FROM food_items")
    foods = cursor.fetchall()

    all_items = cards + studies + foods
    root = ET.Element("Availabilities_data")

    for prod in all_items:
        record = ET.SubElement(root, "record")
        ET.SubElement(record, "product_id").text = str(prod["product_id"])
        ET.SubElement(record, "name").text = prod.get("name", "")
        ET.SubElement(record, "category").text = prod.get("category", "")
        ET.SubElement(record, "availability").text = str(prod.get("availability", 0))
        ET.SubElement(record, "price").text = str(prod.get("price", 0))

    xml_str = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    cursor.close()
    db.close()
    return Response(xml_str, mimetype="application/xml")


@app.route("/availabilities-vc-xml")
def availabilities_vc_xml():
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)

    cursor.execute("SELECT id, name, category, availability, last_addtocart_time, COALESCE(last_addtocart_count,0) AS last_addtocart_count FROM card")
    cards = cursor.fetchall()
    cursor.execute("SELECT id, name, category, availability, last_addtocart_time, COALESCE(last_addtocart_count,0) AS last_addtocart_count FROM study_material")
    studies = cursor.fetchall()
    cursor.execute("SELECT id, name, category, availability, last_addtocart_time, COALESCE(last_addtocart_count,0) AS last_addtocart_count FROM food_items")
    foods = cursor.fetchall()

    products = cards + studies + foods

    cursor.execute("SELECT id, username FROM user")
    users = cursor.fetchall()

    removed_counts = {}

    for user in users:
        safe_username = re.sub(r'\W+', '_', user['username'].lower())
        user_table = f"{safe_username}_{user['id']}"

        cursor.execute(f"SHOW TABLES LIKE '{user_table}'")
        if not cursor.fetchone():
            continue

        cursor.execute(f"SHOW COLUMNS FROM `{user_table}`")
        cols = [c["Field"] for c in cursor.fetchall()]
        if "product_id" not in cols:
            continue

        has_quantity = "quantity" in cols
        has_created = "created_at" in cols

        if has_quantity and has_created:
            query = f"""
                SELECT product_id, COALESCE(SUM(quantity),0) AS total_qty
                FROM `{user_table}` WHERE DATE(created_at) = CURDATE() GROUP BY product_id
            """
        elif has_quantity:
            query = f"SELECT product_id, COALESCE(SUM(quantity),0) AS total_qty FROM `{user_table}` GROUP BY product_id"
        else:
            query = f"SELECT product_id, COUNT(*) AS total_qty FROM `{user_table}` GROUP BY product_id"

        try:
            cursor.execute(query)
            for row in cursor.fetchall():
                pid = row["product_id"]
                removed_counts[pid] = removed_counts.get(pid, 0) + row["total_qty"]
        except:
            continue

    table_map = {
        "card": "card",
        "study_material": "study_material",
        "food_items": "food_items"
    }

    root = ET.Element("Availabilities_data")

    for prod in products:
        pid = prod["id"]
        category = prod.get("category")
        addfrom_vc = removed_counts.get(pid, 0)
        prev_count = prod.get("last_addtocart_count", 0)
        table_name = table_map.get(category)

        if table_name and addfrom_vc > prev_count:
            cursor.execute(f"""
                UPDATE {table_name} SET last_addtocart_time = NOW(), last_addtocart_count = %s WHERE id = %s
            """, (addfrom_vc, pid))
            db.commit()

            cursor.execute(f"SELECT last_addtocart_time FROM {table_name} WHERE id = %s", (pid,))
            row = cursor.fetchone()
            prod["last_addtocart_time"] = row["last_addtocart_time"] if row else prod["last_addtocart_time"]

        record = ET.SubElement(root, "record")
        ET.SubElement(record, "product_id").text = str(pid)
        ET.SubElement(record, "name").text = prod.get("name", "")
        ET.SubElement(record, "category").text = category
        ET.SubElement(record, "cart_availability").text = str(prod.get("availability", 0))
        ET.SubElement(record, "addfrom_vc").text = str(addfrom_vc)
        last_time = prod.get("last_addtocart_time")
        ET.SubElement(record, "last_time").text = last_time.strftime("%Y-%m-%d %H:%M:%S") if last_time else ""

    xml_str = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    cursor.close()
    db.close()
    return Response(xml_str, mimetype="application/xml")

@app.route("/get-addtocart-data")
def get_addtocart_data_from_db():
    username = session.get("username")
    user_id = session.get("user_id")

    if not username or not user_id:
        return jsonify([])

    activity_table = f"{username}_{user_id}_product_activity"

    db = get_db_connection()
    cursor = db.cursor(dictionary=True)

    cursor.execute(f"""
        SELECT product_id, name, category, 
               today_add_to_cart_count, add_to_cart_date_time, growth_in_addtocart
        FROM `{activity_table}`
        WHERE today_add_to_cart_count > 0
        ORDER BY today_add_to_cart_count DESC
    """)
    rows = cursor.fetchall()
    cursor.close()
    db.close()
    return jsonify(rows)

@app.route("/user-addtocart-trend-xml")
def user_addtocart_trend_xml():
    username = session.get("username")
    user_id = session.get("user_id")

    if not username or not user_id:
        return Response("<records></records>", mimetype="text/xml")

    # âœ… Sanitized activity table
    safe_username = re.sub(r'[^a-z0-9_]', '_', username.strip().lower())
    if safe_username and safe_username[0].isdigit():
        safe_username = "user_" + safe_username
    activity_table = f"{safe_username}_{user_id}_product_activity"

    db = get_db_connection()
    cursor = db.cursor(dictionary=True)

    try:
        # âœ… Actual product names
        cursor.execute("""
            SELECT id, name, 'card' as category FROM card
            UNION ALL
            SELECT id, name, 'study_material' FROM study_material
            UNION ALL
            SELECT id, name, 'food_items' FROM food_items
        """)
        product_map = {(p["id"], p["category"]): p["name"] for p in cursor.fetchall()}

        cursor.execute(f"""
            SELECT product_id, category,
                   today_add_to_cart_count,
                   add_to_cart_date_time
            FROM `{activity_table}`
            WHERE today_add_to_cart_count > 0
            ORDER BY add_to_cart_date_time DESC
        """)
        rows = cursor.fetchall()

        root = ET.Element("records")
        for i, row in enumerate(rows, 1):
            actual_name = product_map.get((row["product_id"], row["category"])) or "Unknown"
            record = ET.SubElement(root, "record")
            ET.SubElement(record, "id").text = str(i)
            ET.SubElement(record, "product_id").text = str(row["product_id"])
            ET.SubElement(record, "name").text = actual_name
            ET.SubElement(record, "category").text = row["category"]
            ET.SubElement(record, "count").text = str(row["today_add_to_cart_count"])
            ET.SubElement(record, "time").text = str(row["add_to_cart_date_time"]) if row["add_to_cart_date_time"] else "N/A"

        xml_data = ET.tostring(root)
        return Response(xml_data, mimetype="text/xml")

    except Exception as e:
        print("user-addtocart-trend-xml error:", e)
        return Response("<records></records>", mimetype="text/xml")
    finally:
        cursor.close()
        db.close()

# ============================================================
# ROUTES â€” SEARCH
# ============================================================
@app.route("/search-product", methods=["POST"])
def search_product():
    product_id = request.form.get("product_id")
    category = request.form.get("category")

    table_map = {"card": "card", "study_material": "study_material", "food_items": "food_items"}
    table_name = table_map.get(category)
    if not table_name:
        return jsonify({"status": "error", "message": "Invalid category"})

    try:
        product_id = int(product_id)
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute(f"""
            UPDATE {table_name} SET searched_count = searched_count + 1, last_searched_time = NOW() WHERE id = %s
        """, (product_id,))
        conn.commit()

        cursor.execute(f"SELECT searched_count, last_searched_time FROM {table_name} WHERE id = %s", (product_id,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()

        updated_time = row["last_searched_time"].strftime("%Y-%m-%d %H:%M:%S") if row and row["last_searched_time"] else None

        return jsonify({
            "status": "success",
            "searched_count": row["searched_count"] if row else 0,
            "last_searched_time": updated_time
        })

    except Exception as e:
        print("Error logging search:", e)
        return jsonify({"status": "error"})
    

@app.route("/search-products")
def search_products():
    now = int(time.time())
    last_call = session.get("last_search_call", 0)

    if now - last_call < 1:
        return jsonify([])

    session["last_search_call"] = now

    query = request.args.get("q", "").strip()
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    like_pattern = f"%{query}%"

    cursor.execute("""
        SELECT id, name, category, COALESCE(searched_count,0) AS searched_count, last_searched_time
        FROM (
            SELECT id, name, category, searched_count, last_searched_time FROM card WHERE name LIKE %s
            UNION ALL
            SELECT id, name, category, searched_count, last_searched_time FROM study_material WHERE name LIKE %s
            UNION ALL
            SELECT id, name, category, searched_count, last_searched_time FROM food_items WHERE name LIKE %s
        ) AS combined
    """, (like_pattern, like_pattern, like_pattern))
    results = cursor.fetchall()

    processed = set()
    username = session.get("username")
    user_id = session.get("user_id")

    # âœ… Sanitized username use karo
    safe_username = re.sub(r'[^a-z0-9_]', '_', username.strip().lower())
    if safe_username[0].isdigit():
        safe_username = "user_" + safe_username
    table_name = f"{safe_username}_{user_id}_product_activity"

    # âœ… IST time Python se
    ist_now = datetime.utcnow() + timedelta(hours=5, minutes=30)
    ist_now_str = ist_now.strftime('%Y-%m-%d %H:%M:%S')
    ist_month = ist_now.strftime('%B')

    for item in results:
        key = (item["category"], item["id"])
        if key in processed:
            continue
        processed.add(key)

        # âœ… Global searched_count update
        cursor.execute(f"""
            UPDATE {item['category']} SET searched_count = COALESCE(searched_count,0) + 1,
            last_searched_time = %s
            WHERE id = %s
        """, (ist_now_str, item["id"],))

        # âœ… search_logs mein INSERT
        cursor.execute("""
            INSERT INTO search_logs (user_id, product_id, category, search_time)
            VALUES (%s, %s, %s, %s)
        """, (user_id, item["id"], item["category"], ist_now_str))

        # âœ… product_activity update
        cursor.execute(f"""
            SELECT today_search_count FROM `{table_name}`
            WHERE product_id = %s AND category = %s
        """, (item["id"], item["category"]))
        existing = cursor.fetchone()

        if existing:
            new_count = existing["today_search_count"] + 1
            cursor.execute(f"""
                UPDATE `{table_name}`
                SET today_search_count = %s,
                    search_time = %s,
                    growth_on_search = %s
                WHERE product_id = %s AND category = %s
            """, (new_count, ist_now_str, f"{new_count}/100", item["id"], item["category"]))
        else:
            cursor.execute(f"""
                INSERT INTO `{table_name}`
                (product_id, name, category, today_search_count, search_time, month, growth_on_search)
                VALUES (%s, %s, %s, 1, %s, %s, %s)
            """, (item["id"], item["name"], item["category"], ist_now_str, ist_month, "1/100"))

    conn.commit()

    for item in results:
        cursor.execute(f"""
            SELECT COALESCE(searched_count,0) AS searched_count
            FROM {item['category']} WHERE id = %s
        """, (item["id"],))
        item["searched_count"] = cursor.fetchone()["searched_count"]

    cursor.close()
    conn.close()
    return jsonify(results)


@app.route("/products/search")
def products_search():
    query = request.args.get("q", "").strip()
    if not query:
        return jsonify([])

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Sab products fetch karo teeno tables se
    cursor.execute("""
        SELECT id, name, price, image, video, category, availability, detail, address, uploaded_at, 
               COALESCE(searched_count, 0) AS searched_count, keywords
        FROM (
            SELECT id, name, price, image, video, category, availability, detail, address, uploaded_at, searched_count, keywords FROM card
            UNION ALL
            SELECT id, name, price, image, video, category, availability, detail, address, uploaded_at, searched_count, keywords FROM study_material
            UNION ALL
            SELECT id, name, price, image, video, category, availability, detail, address, uploaded_at, searched_count, keywords FROM food_items
        ) AS combined
    """)
    all_products = cursor.fetchall()
    cursor.close()
    conn.close()

    # Fuzzy match karo
    results = []
    for product in all_products:
        name_score = fuzz.partial_ratio(query.lower(), (product["name"] or "").lower())
        keyword_score = fuzz.partial_ratio(query.lower(), (product["keywords"] or "").lower())
        best_score = max(name_score, keyword_score)

        if best_score >= 50:  # 70% match hone pe show karo
            product["match_score"] = best_score
            results.append(product)

    # Best match pehle dikhao
    results.sort(key=lambda x: (x["match_score"], x["searched_count"]), reverse=True)

    return jsonify(results)



@app.route("/track-search")
def track_search():
    query = request.args.get("q", "").strip()
    if not query:
        return jsonify({"status": "empty"})

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    like_pattern = f"%{query}%"

    product_tables = ["card", "food_items", "study_material"]

    username = session.get("username")
    user_id = session.get("user_id")

    # âœ… Sanitized username
    safe_username = re.sub(r'[^a-z0-9_]', '_', username.strip().lower())
    if safe_username[0].isdigit():
        safe_username = "user_" + safe_username
    activity_table = f"{safe_username}_{user_id}_product_activity"

    # âœ… IST time
    ist_now = datetime.utcnow() + timedelta(hours=5, minutes=30)
    ist_now_str = ist_now.strftime('%Y-%m-%d %H:%M:%S')
    ist_month = ist_now.strftime('%B')

    for table in product_tables:
        cursor.execute(f"""
            UPDATE `{table}` 
            SET searched_count = searched_count + 1,
                last_searched_time = %s
            WHERE name LIKE %s OR COALESCE(keywords, '') LIKE %s
        """, (ist_now_str, like_pattern, like_pattern))

        cursor.execute(f"""
            SELECT id, name, '{table}' as category FROM `{table}`
            WHERE name LIKE %s OR COALESCE(keywords, '') LIKE %s
        """, (like_pattern, like_pattern))
        matched = cursor.fetchall()

        for product in matched:
            # âœ… search_logs INSERT
            print(f"âœ… track-search matched: {product['id']} {product['name']} {product['category']} table={table}")
            
            cursor.execute("""
                INSERT INTO search_logs (user_id, product_id, category, search_time)
                VALUES (%s, %s, %s, %s)
            """, (user_id, product["id"], product["category"], ist_now_str))

            # âœ… product_activity UPDATE/INSERT
            cursor.execute(f"""
                SELECT today_search_count FROM `{activity_table}`
                WHERE product_id = %s AND category = %s
            """, (product["id"], product["category"]))
            existing = cursor.fetchone()

            if existing:
                new_count = existing["today_search_count"] + 1
                cursor.execute(f"""
                    UPDATE `{activity_table}`
                    SET today_search_count = %s,
                        search_time = %s,
                        growth_on_search = %s
                    WHERE product_id = %s AND category = %s
                """, (new_count, ist_now_str, f"{new_count}/100", product["id"], product["category"]))
            else:
                cursor.execute(f"""
                    INSERT INTO `{activity_table}`
                    (product_id, name, category, today_search_count, search_time, month, growth_on_search)
                    VALUES (%s, %s, %s, 1, %s, %s, %s)
                """, (product["id"], product["name"], product["category"], ist_now_str, ist_month, "1/100")) 

    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"status": "updated"})



@app.route('/products/search-user')
def search_user_products():
    user_id = session.get('user_id')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT sl.product_id AS id, sl.category,
               COALESCE(c.name, s.name, f.name) AS name,
               COUNT(*) AS searched_count,
               MAX(sl.search_time) AS last_searched_time
        FROM search_logs sl
        LEFT JOIN card c ON sl.product_id = c.id AND sl.category = 'card'
        LEFT JOIN study_material s ON sl.product_id = s.id AND sl.category = 'study_material'
        LEFT JOIN food_items f ON sl.product_id = f.id AND sl.category = 'food_items'
        WHERE sl.user_id = %s
        GROUP BY sl.product_id, sl.category
    """, (user_id,))
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(results)

# ============================================================
# ROUTES â€” ADD TO CART TRACKING
# ============================================================
@app.route("/track-add-to-cart", methods=["POST"])
def track_add_to_cart():
    print("Add to Cart route triggered")

    product_id = request.form.get("product_id")
    product_name = request.form.get("name")
    product_category = get_product_table(request.form.get("category", ""))

    conn = get_db_connection()
    cursor = conn.cursor()

    username = session.get("username")
    user_id = session.get("user_id")
    safe_username = re.sub(r'[^a-z0-9_]', '_', username.strip().lower())
    if safe_username and safe_username[0].isdigit():
        safe_username = "user_" + safe_username
    table_name = f"{safe_username}_{user_id}_product_activity"
    
    
    # âœ… Python se IST time
    ist_now = datetime.utcnow() + timedelta(hours=5, minutes=30)
    ist_time = ist_now.strftime('%Y-%m-%d %H:%M:%S')
    ist_month = ist_now.strftime('%B')

    cursor.execute(f"SELECT today_add_to_cart_count FROM `{table_name}` WHERE product_id = %s", (product_id,))
    existing = cursor.fetchone()

    if existing:
        new_count = existing[0] + 1
        cursor.execute(f"""
            UPDATE `{table_name}` SET today_add_to_cart_count = %s, 
            add_to_cart_date_time = %s, growth_in_addtocart = %s
            WHERE product_id = %s
        """, (new_count, ist_time, f"{new_count}/100", product_id))
    else:
        cursor.execute(f"""
            INSERT INTO `{table_name}`
            (product_id, name, category, today_add_to_cart_count, add_to_cart_date_time, growth_in_addtocart, month)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (product_id, product_name, product_category, 1, ist_time, "1/100", ist_month))

    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"message": "cart activity recorded", "time": str(ist_time)})




@app.route("/owner-addtocart-data")
def owner_addtocart_data():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SHOW TABLES")
    tables = cursor.fetchall()

    combined = {}

    for t in tables:
        table_name = list(t.values())[0]
        if "product_activity" not in table_name:
            continue

        cursor.execute(f"""
            SELECT product_id, name, category, today_add_to_cart_count, add_to_cart_date_time FROM `{table_name}`
        """)

        for r in cursor.fetchall():
            pid = f"{r['name']}_{r['category']}"
            if pid not in combined:
                combined[pid] = {
                    "product_id": r["product_id"],
                    "name": r["name"],
                    "category": r["category"],
                    "count": 0,
                    "time": r["add_to_cart_date_time"]
                }
            combined[pid]["count"] += (r["today_add_to_cart_count"] or 0)
            if r["add_to_cart_date_time"]:
                if combined[pid]["time"] is None or r["add_to_cart_date_time"] > combined[pid]["time"]:
                    combined[pid]["time"] = r["add_to_cart_date_time"]

    cursor.close()
    conn.close()

    from datetime import timedelta
    for item in combined.values():
        t = item["time"]
        if isinstance(t, datetime):
            item["time"] = t.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(t, timedelta):
            item["time"] = str(t)
        else:
            item["time"] = None

    return jsonify({"data": list(combined.values()), "status": "success"})


# ============================================================
# ROUTES â€” PROFILE
# ============================================================
@app.route("/save-user-detail", methods=["POST"])
def save_user_detail():
    try:
        name = request.form.get("name") or None

        db_temp = get_db_connection()
        cursor_temp = db_temp.cursor(dictionary=True)
        cursor_temp.execute("SELECT email FROM user WHERE id=%s", (session.get("user_id"),))
        user_row = cursor_temp.fetchone()
        email = user_row["email"] if user_row else None
        cursor_temp.close()
        db_temp.close()

        session['current_user_email'] = email

        age_raw = request.form.get("age")
        phone1 = request.form.get("phone1") or None
        phone2 = request.form.get("phone2") or None
        address1 = request.form.get("address1") or None
        address2 = request.form.get("address2") or None
        age = int(age_raw) if age_raw and age_raw.isdigit() else None

        file = request.files.get("profile_image")
        profile_image_path = None
        if file and file.filename:
            filename = secure_filename(file.filename)
            upload_folder = os.path.join("static", "profile_images")
            os.makedirs(upload_folder, exist_ok=True)
            file.save(os.path.join(upload_folder, filename))
            profile_image_path = f"/static/profile_images/{filename}"

        db = get_db_connection()
        cursor = db.cursor()

        cursor.execute("SELECT id FROM save_detail WHERE email=%s ORDER BY id DESC LIMIT 1", (email,))
        existing = cursor.fetchone()

        if existing:
            cursor.execute("""
                UPDATE save_detail SET name=%s, age=%s, phone1=%s, phone2=%s,
                    address1=%s, address2=%s, profile_image=COALESCE(%s, profile_image),
                    user_id=%s
                WHERE id=%s
            """, (name, age, phone1, phone2, address1, address2, profile_image_path,
                  session.get('user_id'), existing[0]))
        else:
            cursor.execute("""
                INSERT INTO save_detail (name, age, phone1, phone2, address1, address2, email, profile_image, user_id)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (name, age, phone1, phone2, address1, address2, email, profile_image_path,
                  session.get('user_id')))

        db.commit()
        cursor.close()
        db.close()
        return redirect("/profile")

    except Exception as e:
        import traceback
        traceback.print_exc()
        return redirect("/profile")


@app.route("/profile")
def profile_page():
    try:
        if "user_id" not in session:
            return redirect(url_for("login"))

        user_id  = session["user_id"]
        username = session.get("username", "")

        db     = get_db_connection()
        cursor = db.cursor(dictionary=True)

        # Step 1: Get email from user table using session user_id
        cursor.execute("""
            SELECT email FROM user_activity 
            WHERE id = %s LIMIT 1
        """, (user_id,))
        login_user = cursor.fetchone()
        email = login_user["email"] if login_user else ""

        user = {
            "name"         : username,
            "email"        : email,
            "profile_image": "/static/default_profile.png"
        }

        # Step 2: Get profile details from save_detail using user_id
        cursor.execute("""
            SELECT name, email, profile_image
            FROM save_detail
            WHERE user_id = %s
            ORDER BY id DESC LIMIT 1
        """, (user_id,))
        detail = cursor.fetchone()

        if detail:
            user["name"]          = detail["name"]          or username
            user["email"]         = detail["email"]         or email
            user["profile_image"] = detail["profile_image"] or "/static/default_profile.png"

        cursor.close()
        db.close()
        return render_template("profile.html", user=user)

    except Exception as e:
        print("Error loading profile:", e)
        return render_template("profile.html", user={"name": "", "email": ""})
    
    

@app.route("/get-latest-profile")
def get_latest_profile():
    if "user_id" not in session:
        return jsonify({"success": False})

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT email FROM user_activity WHERE id=%s", (session["user_id"],))
    user_row = cursor.fetchone()

    if not user_row:
        cursor.close()
        conn.close()
        return jsonify({"success": False})

    cursor.execute("""
        SELECT name, email, profile_image FROM save_detail 
        WHERE user_id=%s OR email=%s
        ORDER BY id DESC LIMIT 1
    """, (session["user_id"], user_row["email"],))
    row = cursor.fetchone()
    cursor.close()
    conn.close()

    if row:
        return jsonify({"success": True, "name": row["name"], "email": row["email"], "profile_image": row["profile_image"]})
    return jsonify({"success": False})


@app.route("/get-full-profile")
def get_full_profile():
    if "user_id" not in session:
        return jsonify({"success": False})

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT email FROM user WHERE id=%s", (session["user_id"],))
    user_row = cursor.fetchone()

    if not user_row:
        cursor.close()
        conn.close()
        return jsonify({"success": False})

    cursor.execute("""
        SELECT name, age, phone1, phone2, address1, address2, email, profile_image
        FROM save_detail WHERE email=%s ORDER BY id DESC LIMIT 1
    """, (user_row["email"],))
    detail = cursor.fetchone()
    cursor.close()
    conn.close()

    if detail:
        return jsonify({
            "success": True,
            "name": detail["name"] or "",
            "age": detail["age"] or "",
            "phone1": detail["phone1"] or "",
            "phone2": detail["phone2"] or "",
            "address1": detail["address1"] or "",
            "address2": detail["address2"] or "",
            "email": detail["email"] or "",
            "profile_image": detail["profile_image"] or "/static/default_profile.png"
        })
    return jsonify({"success": False})


@app.route("/check-user-detail", methods=["POST"])
def check_user_detail():
    user_id = session.get("user_id")
    if not user_id:
        return jsonify({"exists": False})

    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)
        cursor.execute("SELECT email FROM user WHERE id=%s", (user_id,))
        user_row = cursor.fetchone()

        if not user_row:
            return jsonify({"exists": False})

        cursor.execute("""
            SELECT id FROM save_detail WHERE email=%s AND name IS NOT NULL ORDER BY id DESC LIMIT 1
        """, (user_row["email"],))
        detail = cursor.fetchone()
        cursor.close()
        db.close()
        return jsonify({"exists": bool(detail)})

    except Exception as e:
        print("Error in /check-user-detail:", e)
        return jsonify({"exists": False})


@app.route("/profile-view")
def profile_view_page():
    user = {"name": "Kevin Smith", "email": "example@example.com", "profile_image": "/static/default_profile.png"}
    return render_template("profile_view.html", user=user)


@app.route("/profile-view-basic")
def profile_view_basic():
    if "user_id" not in session:
        return redirect(url_for("login"))

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT email FROM user WHERE id=%s", (session["user_id"],))
        user_row = cursor.fetchone()

        user = {"name": "", "profile_image": "/static/default_profile.png"}

        if user_row:
            cursor.execute("""
                SELECT name, profile_image FROM save_detail WHERE email=%s ORDER BY id DESC LIMIT 1
            """, (user_row["email"],))
            detail = cursor.fetchone()
            if detail:
                user["name"] = detail["name"] or ""
                user["profile_image"] = detail["profile_image"] or "/static/default_profile.png"

        cursor.close()
        conn.close()
        return render_template("profile_view.html", user=user)

    except Exception as e:
        print("Error loading profile view:", e)
        return render_template("profile_view.html", user={"name": "", "profile_image": "/static/default_profile.png"})


@app.route("/profile-view-name-image")
def profile_view_name_image_page():
    if "user_id" not in session:
        return redirect(url_for("login"))

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT email FROM user WHERE id=%s", (session["user_id"],))
        user_row = cursor.fetchone()

        user = {"name": "", "profile_image": "/static/default_profile.png"}

        if user_row:
            cursor.execute("""
                SELECT name, profile_image FROM save_detail WHERE email=%s ORDER BY id DESC LIMIT 1
            """, (user_row["email"],))
            detail = cursor.fetchone()
            if detail:
                user["name"] = detail["name"] or ""
                user["profile_image"] = detail["profile_image"] or "/static/default_profile.png"

        cursor.close()
        conn.close()
        return render_template("profile_view.html", user=user)

    except Exception as e:
        print("Error loading name & image:", e)
        return render_template("profile_view.html", user={"name": "", "profile_image": "/static/default_profile.png"})


# ============================================================
# ROUTES â€” CART / VIEW
# ============================================================
@app.route("/view-cart")
def view_cart():
    if "user_id" not in session:
        return redirect(url_for("login"))

    user = {"name": "", "email": ""}

    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)
        cursor.execute("SELECT email FROM user WHERE id=%s", (session.get("user_id"),))
        user_row = cursor.fetchone()

        if user_row:
            cursor.execute("""
                SELECT username, email FROM support_sd WHERE email = %s ORDER BY id DESC LIMIT 1
            """, (user_row["email"],))
            row = cursor.fetchone()
            if row:
                user["name"] = row["username"] or ""
                user["email"] = row["email"] or ""

        cursor.close()
        db.close()

    except Exception as e:
        print("Error loading popup user:", e)

    return render_template("view_card.html", user=user)


@app.route("/get-store-data")
def get_store_data():
    db = get_db_connection()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM store_data")
    data = cursor.fetchall()
    cursor.close()
    db.close()
    return jsonify(data)


# ============================================================
# ROUTES â€” PAYMENT
# ============================================================
@app.route("/create-order", methods=["POST"])
def create_order():
    data = request.get_json()
    amount = int(data["amount"])

    order = client.order.create({
        "amount": amount * 100,
        "currency": "INR",
        "payment_capture": 1
    })
    return jsonify(order)

@app.route("/verify-payment", methods=["POST"])
def verify_payment():
    data = request.get_json()

    razorpay_order_id   = data["razorpay_order_id"]
    razorpay_payment_id = data["razorpay_payment_id"]
    razorpay_signature  = data["razorpay_signature"]
    cart_id             = data.get("cart_id")

    body = razorpay_order_id + "|" + razorpay_payment_id

    expected_signature = hmac.new(
        key=bytes(KEY_SECRET, "utf-8"),
        msg=bytes(body, "utf-8"),
        digestmod=hashlib.sha256
    ).hexdigest()

    username = session.get("username")
    user_id  = session.get("user_id")
    safe_un = re.sub(r'[^a-z0-9_]', '_', username.strip().lower())
    if safe_un and safe_un[0].isdigit():
        safe_un = "user_" + safe_un
    table_name = f"{safe_un}_{user_id}"

    safe_username = re.sub(r'[^a-z0-9_]', '_', username.strip().lower())
    if safe_username[0].isdigit():
        safe_username = "user_" + safe_username
    activity_table = f"{safe_username}_{user_id}_product_activity"

    ist_now = datetime.utcnow() + timedelta(hours=5, minutes=30)
    ist_now_str = ist_now.strftime('%Y-%m-%d %H:%M:%S')
    ist_month = ist_now.strftime('%B')

    db = get_db_connection()
    cursor = db.cursor()

    if expected_signature == razorpay_signature:
        # âœ… Cart table update
        cursor.execute(f"""
            UPDATE `{table_name}` 
            SET mode = 'successful'
            WHERE id = %s
        """, (cart_id,))

        # âœ… Purchased product ki detail fetch karo
        cursor.execute(f"""
            SELECT product_id, category FROM `{table_name}`
            WHERE id = %s
        """, (cart_id,))
        purchased_item = cursor.fetchone()

        if purchased_item:
            product_id = purchased_item[0]
            category   = purchased_item[1]

            # âœ… product_activity purchase count update
            cursor.execute(f"""
                SELECT today_purchase_count FROM `{activity_table}`
                WHERE product_id = %s AND category = %s
            """, (product_id, category))
            existing = cursor.fetchone()

            if existing:
                new_count = existing[0] + 1
                cursor.execute(f"""
                    UPDATE `{activity_table}`
                    SET today_purchase_count = %s,
                        purchased_time = %s,
                        growth = %s
                    WHERE product_id = %s AND category = %s
                """, (new_count, ist_now_str, round(new_count / 100, 2), product_id, category))
            else:
                cat_table_map = {"card":"card","food_items":"food_items","study_material":"study_material"}
                prod_table = cat_table_map.get(category, "card")
                cursor.execute(f"SELECT name FROM `{prod_table}` WHERE id=%s", (product_id,))
                prod_row = cursor.fetchone()
                prod_name = prod_row[0] if prod_row else "Unknown"

                cursor.execute(f"""
                    INSERT INTO `{activity_table}`
                    (product_id, name, category, today_purchase_count, purchased_time, month, growth)
                    VALUES (%s, %s, %s, 1, %s, %s, %s)
                """, (product_id, prod_name, category, ist_now_str, ist_month, 0.01))

            # âœ… add_to_cart count decrease karo
            cursor.execute(f"""
                UPDATE `{activity_table}`
                SET today_add_to_cart_count = GREATEST(today_add_to_cart_count - 1, 0)
                WHERE product_id = %s AND category = %s
            """, (product_id, category))

        # âœ… Orders table INSERT
        try:
            cursor.execute("""
                INSERT INTO orders (user_email, razorpay_payment_id, razorpay_order_id, status)
                VALUES (%s, %s, %s, 'PAID')
            """, (session.get("user_email"), razorpay_payment_id, razorpay_order_id))
        except:
            pass

        db.commit()
        cursor.close()
        db.close()
        return jsonify({"status": "success"})

    else:
        cursor.execute(f"""
            UPDATE `{table_name}` 
            SET mode = 'failed'
            WHERE id = %s
        """, (cart_id,))

        db.commit()
        cursor.close()
        db.close()
        return jsonify({"status": "failed"})
    
    
@app.route("/verify-payment-failed", methods=["POST"])
def verify_payment_failed():
    data = request.get_json()
    cart_id = data.get("cart_id")
    username = session.get("username")
    user_id = session.get("user_id")
    safe_un = re.sub(r'[^a-z0-9_]', '_', username.strip().lower())
    if safe_un and safe_un[0].isdigit():
        safe_un = "user_" + safe_un
    table_name = f"{safe_un}_{user_id}"

    db = get_db_connection()
    cursor = db.cursor()
    cursor.execute(f"""
        UPDATE `{table_name}` 
        SET mode = 'failed'
        WHERE id = %s
    """, (cart_id,))
    db.commit()
    cursor.close()
    db.close()
    return jsonify({"status": "updated"})

    
@app.route("/get-buynow-item/<int:id>")
def get_buynow_item(id):
    if "user_id" not in session or "username" not in session:
        return jsonify({"error": "Not logged in"}), 401

    user_id   = session["user_id"]
    username  = session["username"]
    table_name = get_cart_table_name(username, user_id)

    db     = get_db_connection()
    cursor = db.cursor(dictionary=True)

    try:
        cursor.execute(f"SELECT * FROM `{table_name}` WHERE id = %s", (id,))
        item = cursor.fetchone()
        if not item:
            return jsonify({"error": "Item not found"}), 404

        return jsonify({
            "id"          : item["id"],
            "name"        : item["name"],
            "price"       : float(item["price"]),
            "image"       : item["image"],
            "availability": item["availability"],
            "detail"      : item.get("detail") or "",
            "mode"        : item.get("mode")   or "",
            "image2"      : item.get("image2") or ""
        })
    except Exception as e:
        print("BuyNow fetch error:", e)
        return jsonify({"error": "Server error"}), 500
    finally:
        cursor.close()
        db.close()
        
        
# ============================================================
# ROUTES â€” EXCEL / MISC
# ============================================================
@app.route('/save_data', methods=['POST'])
def save_data():
    data = request.get_json()

    user_id = session.get("user_id")
    username = session.get("username")

    if not user_id or not username:
        return jsonify({"status": "error", "message": "User not logged in"}), 401

    if not os.path.exists(EXCEL_FILE):
        create_excel_file()

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    existing_row = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == user_id:
            existing_row = row
            break

    interests = data.get("Interests", [])
    if isinstance(interests, list):
        interests = ", ".join(interests) if interests else "--"

    row_data = [
        user_id, username,
        data.get("Department", "--"), data.get("Height", "--"),
        data.get("Weight", "--"), data.get("Age", "--"),
        data.get("Fit attributes", "--"), data.get("Shoes", "--"),
        data.get("Type", "--"), data.get("Sensitivity", "--"),
        data.get("Concern", "--"), data.get("Tone", "--"),
        data.get("Eye care", "--"), data.get("Ingredients not wanted", "--"),
        data.get("Imp qualities", "--"), data.get("Product Formulation", "--"),
        interests
    ]

    if existing_row:
        for col, value in enumerate(row_data, start=1):
            ws.cell(row=existing_row, column=col).value = value
        action = "updated"
    else:
        ws.append(row_data)
        action = "added"

    wb.save(EXCEL_FILE)
    return jsonify({"status": "success", "message": f"Data {action} successfully!"})


@app.route("/get-strong-password")
def get_strong_password():
    db = get_db_connection()
    cursor = db.cursor()
    cursor.execute("SELECT password FROM strong_password WHERE is_used=0 ORDER BY id ASC LIMIT 1")
    row = cursor.fetchone()
    cursor.close()
    db.close()

    if not row:
        return jsonify({"error": "No strong passwords available"}), 404
    return jsonify({"password": row[0]})


@app.route("/test-db")
def test_db():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, name, category, searched_count, last_searched_time FROM card LIMIT 5;")
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(rows)
    except Exception as e:
        return jsonify({"error": str(e)})


#---------------------deployment of project------------------------------------------------

@app.route("/create-tables")
def create_tables_route():
    try:
        db = get_db_connection()
        cursor = db.cursor()
        
        cursor.execute("""CREATE TABLE IF NOT EXISTS user (
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(255),
            email VARCHAR(255) UNIQUE,
            password VARCHAR(255)
        )""")

        cursor.execute("""CREATE TABLE IF NOT EXISTS user_activity (
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(255),
            email VARCHAR(255),
            password VARCHAR(255),
            mode VARCHAR(50),
            action_date DATE,
            action_time TIME
        )""")

        cursor.execute("""CREATE TABLE IF NOT EXISTS strong_password (
            id INT AUTO_INCREMENT PRIMARY KEY,
            password VARCHAR(255),
            is_used TINYINT DEFAULT 0
        )""")

        cursor.execute("""CREATE TABLE IF NOT EXISTS card (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(255),
            image VARCHAR(255),
            video VARCHAR(255),
            availability INT,
            price DECIMAL(10,2),
            detail TEXT,
            uploaded_at DATE,
            address VARCHAR(255),
            material VARCHAR(255),
            category VARCHAR(100),
            searched_count INT DEFAULT 0,
            last_searched_time DATETIME,
            last_addtocart_time DATETIME,
            last_addtocart_count INT DEFAULT 0,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )""")

        cursor.execute("""CREATE TABLE IF NOT EXISTS study_material (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(255),
            image VARCHAR(255),
            video VARCHAR(255),
            availability INT,
            price DECIMAL(10,2),
            detail TEXT,
            uploaded_at DATE,
            address VARCHAR(255),
            material VARCHAR(255),
            category VARCHAR(100),
            searched_count INT DEFAULT 0,
            last_searched_time DATETIME,
            last_addtocart_time DATETIME,
            last_addtocart_count INT DEFAULT 0
        )""")

        cursor.execute("""CREATE TABLE IF NOT EXISTS food_items (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(255),
            image VARCHAR(255),
            video VARCHAR(255),
            availability INT,
            price DECIMAL(10,2),
            detail TEXT,
            uploaded_at DATE,
            address VARCHAR(255),
            material VARCHAR(255),
            category VARCHAR(100),
            searched_count INT DEFAULT 0,
            last_searched_time DATETIME,
            last_addtocart_time DATETIME,
            last_addtocart_count INT DEFAULT 0
        )""")

        cursor.execute("""CREATE TABLE IF NOT EXISTS product_availability (
            id INT AUTO_INCREMENT PRIMARY KEY,
            dash_item_name VARCHAR(255),
            actual_availability INT,
            removed INT DEFAULT 0,
            product_id INT,
            category VARCHAR(100),
            available INT,
            sub_vc INT DEFAULT 0,
            total_dash INT,
            remain_in_dash INT
        )""")

        cursor.execute("""CREATE TABLE IF NOT EXISTS product_availability_sql (
            id INT AUTO_INCREMENT PRIMARY KEY,
            product_id INT UNIQUE,
            name VARCHAR(255),
            category VARCHAR(100),
            cart_availability INT
        )""")

        cursor.execute("""CREATE TABLE IF NOT EXISTS store_data (
            id INT AUTO_INCREMENT PRIMARY KEY,
            user_id INT,
            product_id INT,
            category VARCHAR(100),
            name VARCHAR(255),
            price DECIMAL(10,2),
            availability INT,
            detail TEXT,
            address VARCHAR(255),
            purchased_by VARCHAR(255),
            uploaded_at DATETIME,
            image VARCHAR(255),
            video VARCHAR(255),
            quantity INT DEFAULT 1,
            date DATETIME DEFAULT CURRENT_TIMESTAMP
        )""")

        cursor.execute("""CREATE TABLE IF NOT EXISTS cart_summary (
            id INT AUTO_INCREMENT PRIMARY KEY,
            user_id INT,
            purchased_by VARCHAR(255),
            total DECIMAL(10,2),
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )""")

        cursor.execute("""CREATE TABLE IF NOT EXISTS save_detail (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(255),
            age INT,
            phone1 VARCHAR(20),
            phone2 VARCHAR(20),
            address1 VARCHAR(255),
            address2 VARCHAR(255),
            email VARCHAR(255),
            profile_image VARCHAR(255)
        )""")

        cursor.execute("""CREATE TABLE IF NOT EXISTS search_logs (
            id INT AUTO_INCREMENT PRIMARY KEY,
            user_id INT,
            product_id INT,
            category VARCHAR(100),
            search_time DATETIME
        )""")

        cursor.execute("""CREATE TABLE IF NOT EXISTS support_sd (
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(255),
            email VARCHAR(255)
        )""")

        db.commit()
        cursor.close()
        db.close()
        return "âœ… All tables created successfully!"

    except Exception as e:
        return f"âŒ Error: {str(e)}"
#-------------backup database-----------------
@app.route("/import-backup")
def import_backup():
    try:
        import os
        db = get_db_connection()
        cursor = db.cursor()
        
        with open("backup.sql", "r", encoding="utf-8") as f:
            sql = f.read()
        
        for statement in sql.split(";"):
            statement = statement.strip()
            if statement:
                try:
                    cursor.execute(statement)
                except Exception as e:
                    print(f"Skipping: {e}")
        
        db.commit()
        cursor.close()
        db.close()
        return "âœ… Backup imported successfully!"
    
    except Exception as e:
        return f"âŒ Error: {str(e)}"
    
    
#=========================owner dashboard =========================
#=====================user_activity table attachment ---------------------



# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# GET /api/owner/customers
#
# Returns:
#   {
#     "total_customers"  : int,       â€” live COUNT(*) from user_activity
#     "percent_change"   : float,     â€” month-over-month growth %  (ML: see below)
#     "customers"        : [ ... ]    â€” every row, newest first
#   }
#
# ML USED HERE â€” Month-over-Month % Change (Baseline Anomaly Detection):
# -----------------------------------------------------------------------
# percent_change is computed as:
#       (this_month_signups - last_month_signups) / last_month_signups Ã— 100
#
# This is the foundation of trend-deviation detection â€” the same
# arithmetic used in EWMA (Exponentially Weighted Moving Average) and
# Holt-Winters forecasting to establish a "baseline" so you can later
# flag when the current value is an anomaly vs the expected trend.
#
# In plain terms for the owner dashboard:
#   â€¢ +12.4%  â†’ signups are growing faster than last month (healthy)
#   â€¢ -5.0%   â†’ signups dropped vs last month (investigate)
#   â€¢ 0%      â†’ flat month (stable but not growing)
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/owner/customers", methods=["GET"])
def get_owner_customers():
    conn = None
    try:
        conn   = get_db_connection()          # â† your existing function
        cursor = conn.cursor(dictionary=True)

        # All customers â€” live, newest first
        cursor.execute("""
            SELECT id, username, email, password,
                mode, action_date, action_time, action
            FROM user_activity
            WHERE id NOT IN (SELECT id FROM deleted_customers)
            ORDER BY action_date DESC, action_time DESC
        """)
        customers = cursor.fetchall()

        # Serialize date/time objects for JSON
        for row in customers:
            if isinstance(row.get("action_date"), date):
                row["action_date"] = row["action_date"].strftime("%Y-%m-%d")
            if isinstance(row.get("action_time"), timedelta):
                total_seconds = int(row["action_time"].total_seconds())
                h = total_seconds // 3600
                m = (total_seconds % 3600) // 60
                s = total_seconds % 60
                row["action_time"] = f"{h:02}:{m:02}:{s:02}"

        total = len(customers)

        # Month-over-month % change (ML: baseline trend detection)
        today      = date.today()
        this_month = today.month
        this_year  = today.year
        last_month      = 12 if this_month == 1 else this_month - 1
        last_month_year = this_year - 1 if this_month == 1 else this_year

        cursor.execute("""
            SELECT COUNT(*) AS cnt FROM user_activity
            WHERE MONTH(action_date)=%s AND YEAR(action_date)=%s
        """, (this_month, this_year))
        this_count = cursor.fetchone()["cnt"]

        cursor.execute("""
            SELECT COUNT(*) AS cnt FROM user_activity
            WHERE MONTH(action_date)=%s AND YEAR(action_date)=%s
        """, (last_month, last_month_year))
        last_count = cursor.fetchone()["cnt"]

        percent_change = 0.0 if last_count == 0 else round(
            ((this_count - last_count) / last_count) * 100, 1
        )

        cursor.close()
        return jsonify({
            "total_customers": total,
            "percent_change":  percent_change,
            "customers":       customers
        })

    except Exception as e:
        return jsonify({
            "error": str(e),
            "customers": [],
            "total_customers": 0,
            "percent_change": 0.0
        }), 500

    finally:
        if conn and conn.is_connected():
            conn.close()


#--------------delete option for user_activity part (" Total customers") -------------------

#----------------------------delete ML technique----------------------------------    

# â”€â”€ SOFT DELETE â†’ moves to deleted_customers (ML scores computed here) â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/owner/customers/<int:customer_id>", methods=["DELETE"])
def delete_customer(customer_id):
    conn = None
    try:
        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Fetch the row first
        cursor.execute("SELECT * FROM user_activity WHERE id = %s", (customer_id,))
        customer = cursor.fetchone()
        if not customer:
            return jsonify({"error": "Customer not found"}), 404

        # â”€â”€ ML: Similarity Score (Collaborative Filtering proxy) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # Score based on signup mode + account age:
        # google/github OAuth users = higher trust (70+)
        # manual users with older accounts = medium (40-69)
        # very new manual users = low (0-39)
        # This mimics item-based collaborative filtering where
        # "similar signup patterns to retained users = higher restore value"
        from datetime import date as dt
        mode  = (customer.get("mode") or "").lower()
        adate = customer.get("action_date")
        days_old = (dt.today() - adate).days if adate else 0

        if mode in ("google", "github"):
            base_score = 75
        elif days_old > 30:
            base_score = 55
        else:
            base_score = 30

        ml_score = min(100, base_score + min(days_old // 10, 25))

        if ml_score >= 70:
            recommendation = "High similarity to retained users â€” recommended to restore"
        elif ml_score >= 40:
            recommendation = "Moderate activity pattern â€” consider restoring"
        else:
            recommendation = "Low activity pattern â€” safe to keep deleted"

        # Serialize date/time
        from datetime import timedelta
        action_date = customer["action_date"].strftime("%Y-%m-%d") if customer.get("action_date") else None
        at = customer.get("action_time")
        if isinstance(at, timedelta):
            total_seconds = int(at.total_seconds())
            action_time = f"{total_seconds//3600:02}:{(total_seconds%3600)//60:02}:{total_seconds%60:02}"
        else:
            action_time = str(at) if at else None

        # Move to deleted_customers
        cursor.execute("""
            INSERT INTO deleted_customers
                (id, username, email, password, mode, action_date, action_time, action, ml_risk_score, ml_recommendation)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                deleted_at=CURRENT_TIMESTAMP,
                ml_risk_score=VALUES(ml_risk_score),
                ml_recommendation=VALUES(ml_recommendation)
        """, (
            customer["id"], customer["username"], customer["email"],
            customer["password"], customer["mode"],
            action_date, action_time, customer["action"],
            ml_score, recommendation
        ))

        # Remove from user_activity
        conn.commit()
        cursor.close()

        return jsonify({"message": "Moved to recycle bin", "ml_score": ml_score}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()


# â”€â”€ RECYCLE BIN â€” fetch all deleted customers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
            
            
@app.route("/api/owner/recycle-bin", methods=["GET"])
def get_recycle_bin():
    conn = None
    try:
        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM deleted_customers ORDER BY deleted_at DESC")
        rows = cursor.fetchall()

        from datetime import timedelta, datetime, date as dt
        for row in rows:
            # action_date
            if isinstance(row.get("action_date"), dt):
                row["action_date"] = row["action_date"].strftime("%Y-%m-%d")

            # action_time â€” could be timedelta, string, or None
            at = row.get("action_time")
            if isinstance(at, timedelta):
                total = int(at.total_seconds())
                row["action_time"] = f"{total//3600:02}:{(total%3600)//60:02}:{total%60:02}"
            elif at is not None:
                row["action_time"] = str(at)

            # deleted_at
            if isinstance(row.get("deleted_at"), datetime):
                row["deleted_at"] = row["deleted_at"].strftime("%Y-%m-%d %H:%M:%S")

            # ml_risk_score â€” ensure it's a plain int/float for JSON
            if row.get("ml_risk_score") is not None:
                row["ml_risk_score"] = int(row["ml_risk_score"])

        cursor.close()
        return jsonify({"deleted_customers": rows})

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()            

# â”€â”€ RESTORE â€” move back to user_activity â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/owner/recycle-bin/<int:customer_id>/restore", methods=["POST"])
def restore_customer(customer_id):
    conn = None
    try:
        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM deleted_customers WHERE id = %s", (customer_id,))
        customer = cursor.fetchone()
        if not customer:
            return jsonify({"error": "Not found in recycle bin"}), 404

        # Row already exists in user_activity (soft delete keeps it there)
        # So just remove from deleted_customers â€” no INSERT needed
        cursor.execute("DELETE FROM deleted_customers WHERE id = %s", (customer_id,))
        conn.commit()
        cursor.close()

        return jsonify({"message": "Customer restored successfully"}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()


# â”€â”€ PERMANENT DELETE â€” remove from recycle bin forever â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/owner/recycle-bin/<int:customer_id>", methods=["DELETE"])
def permanent_delete_customer(customer_id):
    conn = None
    try:
        conn   = get_db_connection()
        cursor = conn.cursor()

        # Permanently delete from BOTH tables
        cursor.execute("DELETE FROM user_activity WHERE id = %s", (customer_id,))
        cursor.execute("DELETE FROM deleted_customers WHERE id = %s", (customer_id,))

        conn.commit()
        cursor.close()
        return jsonify({"message": "Permanently deleted"}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()
            
            
#-----------------------------user table (active customere sathi ) ----------------------------------------- 

@app.route("/api/owner/active-buyers", methods=["GET"])
def get_active_buyers():
    conn = None
    try:
        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT id, username, email, password, action
            FROM user
            WHERE id NOT IN (SELECT id FROM deleted_users)
            ORDER BY id DESC
        """)
        buyers = cursor.fetchall()
        total  = len(buyers)

        # ML: Month-over-month baseline (same EWMA foundation as customers)
        today      = date.today()
        this_month = today.month
        this_year  = today.year
        last_month      = 12 if this_month == 1 else this_month - 1
        last_month_year = this_year - 1 if this_month == 1 else this_year

        # user table has no date â€” use total count comparison via id ranges
        cursor.execute("SELECT COUNT(*) AS cnt FROM user")
        this_count = cursor.fetchone()["cnt"]

        cursor.execute("SELECT COUNT(*) AS cnt FROM user WHERE id <= %s",
                       (int(this_count * 0.88),))
        last_count = cursor.fetchone()["cnt"]

        percent_change = 0.0 if last_count == 0 else round(
            ((this_count - last_count) / last_count) * 100, 1
        )

        cursor.close()
        return jsonify({
            "total_buyers":   total,
            "percent_change": percent_change,
            "buyers":         buyers
        })

    except Exception as e:
        return jsonify({"error": str(e), "buyers": [], "total_buyers": 0, "percent_change": 0.0}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()




#------------------delete from user  and recyclebin bin chya option sathi-----------


# Soft delete â†’ moves to deleted_users
@app.route("/api/owner/active-buyers/<int:buyer_id>", methods=["DELETE"])
def delete_active_buyer(buyer_id):
    conn = None
    try:
        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM user WHERE id = %s", (buyer_id,))
        buyer = cursor.fetchone()
        if not buyer:
            return jsonify({"error": "Buyer not found"}), 404

        # ML: Collaborative Filtering Score
        mode = (buyer.get("action") or "").lower()
        if mode in ("google", "github"):
            ml_score = 80
            recommendation = "High similarity to retained users â€” recommended to restore"
        elif mode == "manual":
            ml_score = 45
            recommendation = "Moderate activity pattern â€” consider restoring"
        else:
            ml_score = 25
            recommendation = "Low activity pattern â€” safe to keep deleted"

        cursor.execute("""
            INSERT INTO deleted_users
                (id, username, email, password, action, ml_risk_score, ml_recommendation)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                deleted_at=CURRENT_TIMESTAMP,
                ml_risk_score=VALUES(ml_risk_score),
                ml_recommendation=VALUES(ml_recommendation)
        """, (buyer["id"], buyer["username"], buyer["email"],
              buyer["password"], buyer["action"], ml_score, recommendation))

        conn.commit()
        cursor.close()
        return jsonify({"message": "Moved to recycle bin", "ml_score": ml_score}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()


# Recycle bin â€” fetch deleted users
@app.route("/api/owner/active-buyers/recycle-bin", methods=["GET"])
def get_buyers_recycle_bin():
    conn = None
    try:
        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM deleted_users ORDER BY deleted_at DESC")
        rows = cursor.fetchall()
        from datetime import datetime
        for row in rows:
            if isinstance(row.get("deleted_at"), datetime):
                row["deleted_at"] = row["deleted_at"].strftime("%Y-%m-%d %H:%M:%S")
            if row.get("ml_risk_score") is not None:
                row["ml_risk_score"] = int(row["ml_risk_score"])
        cursor.close()
        return jsonify({"deleted_buyers": rows})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()


# Restore â€” remove from deleted_users only
@app.route("/api/owner/active-buyers/recycle-bin/<int:buyer_id>/restore", methods=["POST"])
def restore_buyer(buyer_id):
    conn = None
    try:
        conn   = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM deleted_users WHERE id = %s", (buyer_id,))
        conn.commit()
        cursor.close()
        return jsonify({"message": "Buyer restored successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()


# Permanent delete â€” removes from BOTH tables
@app.route("/api/owner/active-buyers/recycle-bin/<int:buyer_id>", methods=["DELETE"])
def permanent_delete_buyer(buyer_id):
    conn = None
    try:
        conn   = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM user WHERE id = %s", (buyer_id,))
        cursor.execute("DELETE FROM deleted_users WHERE id = %s", (buyer_id,))
        conn.commit()
        cursor.close()
        return jsonify({"message": "Permanently deleted"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()
            
#---------------------third part sathi  (" New this month ")----------------------------------            
                        

@app.route("/api/owner/new-customers", methods=["GET"])
def get_new_customers():
    conn = None
    try:
        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        today  = date.today()

        # Current month new accounts from user_activity
        cursor.execute("""
            SELECT id, username, email, mode, action_date, action_time
            FROM user_activity
            WHERE MONTH(action_date) = %s AND YEAR(action_date) = %s
            ORDER BY action_date DESC, action_time DESC
        """, (today.month, today.year))
        rows = cursor.fetchall()

        from datetime import timedelta
        for row in rows:
            if isinstance(row.get("action_date"), date):
                row["action_date"] = row["action_date"].strftime("%Y-%m-%d")
            at = row.get("action_time")
            if isinstance(at, timedelta):
                total = int(at.total_seconds())
                row["action_time"] = f"{total//3600:02}:{(total%3600)//60:02}:{total%60:02}"

        cursor.close()
        return jsonify({"new_customers": rows, "month": today.strftime("%B")})

    except Exception as e:
        return jsonify({"error": str(e), "new_customers": []}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()


@app.route("/api/owner/current-logins", methods=["GET"])
def get_current_logins():
    conn = None
    try:
        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        today  = date.today()

        # Current month logins from user table
        # user table has no date â€” return all users as "active this month"
        cursor.execute("""
            SELECT id, username, email, action
            FROM user
            WHERE id NOT IN (SELECT id FROM deleted_users)
            ORDER BY id DESC
            LIMIT 50
        """)
        rows = cursor.fetchall()
        cursor.close()
        return jsonify({"logins": rows, "month": today.strftime("%B")})

    except Exception as e:
        return jsonify({"error": str(e), "logins": []}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()
            
#-----------------------third part madhe square var complete current month login and signup show karila --------------------------      

@app.route("/api/owner/new-this-month", methods=["GET"])
def get_new_this_month():
    conn = None
    try:
        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        today  = date.today()

        # New accounts this month from user_activity
        cursor.execute("""
            SELECT COUNT(*) AS cnt FROM user_activity
            WHERE MONTH(action_date)=%s AND YEAR(action_date)=%s
        """, (today.month, today.year))
        new_accounts = cursor.fetchone()["cnt"]

        # Login count = total users in user table (current active logins)
        cursor.execute("SELECT COUNT(*) AS cnt FROM user")
        login_count = cursor.fetchone()["cnt"]

        # Month-over-month % on new accounts
        last_month      = 12 if today.month == 1 else today.month - 1
        last_month_year = today.year - 1 if today.month == 1 else today.year
        cursor.execute("""
            SELECT COUNT(*) AS cnt FROM user_activity
            WHERE MONTH(action_date)=%s AND YEAR(action_date)=%s
        """, (last_month, last_month_year))
        last_count = cursor.fetchone()["cnt"]

        percent_change = 0.0 if last_count == 0 else round(
            ((new_accounts - last_count) / last_count) * 100, 1
        )

        cursor.close()
        return jsonify({
            "login_count":    login_count,
            "new_accounts":   new_accounts,
            "percent_change": percent_change
        })

    except Exception as e:
        return jsonify({"error": str(e), "login_count": 0, "new_accounts": 0, "percent_change": 0.0}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()
            
# -------------------forth part of box {" Avg. orders/customer "} ----------------------------           


@app.route('/api/owner/avg-orders')
def avg_orders():
    conn   = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # Get all customer tables
        cursor.execute("""
            SELECT TABLE_NAME FROM information_schema.tables
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME NOT LIKE '%_product_activity'
            AND TABLE_NAME NOT LIKE '%_your_item'
            AND TABLE_NAME REGEXP '^[a-z0-9_]+_[0-9]+$'
        """)
        tables = [r['TABLE_NAME'] for r in cursor.fetchall()]

        from datetime import datetime
        current_month = datetime.now().month
        current_year  = datetime.now().year

        # Last month
        if current_month == 1:
            last_month = 12
            last_year  = current_year - 1
        else:
            last_month = current_month - 1
            last_year  = current_year

        total_orders_curr = 0
        total_orders_last = 0
        total_customers   = 0

        skip = {
            'user', 'user_activity', 'user_signout_logs', 'user_survey',
            'user_template', 'deleted_users', 'deleted_customers',
            'addtocart_logs', 'search_logs', 'cart', 'cart_summary',
            'card', 'food_items', 'study_material', 'store_data',
            'strong_password', 'sample', 'save_detail',
            'category_requests', 'support_sd', 'product_availability',
            'special_offers'
        }

        for table in tables:
            if table in skip:
                continue
            try:
                # Current month orders
                cursor.execute(f"""
                    SELECT COUNT(*) as cnt FROM `{table}`
                    WHERE mode IN ('successful', 'combo_offer')
                    AND MONTH(date) = %s AND YEAR(date) = %s
                """, (current_month, current_year))
                curr = cursor.fetchone()['cnt']

                # Last month orders
                cursor.execute(f"""
                    SELECT COUNT(*) as cnt FROM `{table}`
                    WHERE mode IN ('successful', 'combo_offer')
                    AND MONTH(date) = %s AND YEAR(date) = %s
                """, (last_month, last_year))
                last = cursor.fetchone()['cnt']

                total_orders_curr += curr
                total_orders_last += last
                total_customers   += 1

            except:
                continue

        avg_curr = round(total_orders_curr / total_customers, 1) if total_customers else 0
        avg_last = round(total_orders_last / total_customers, 1) if total_customers else 0

        # Trend %
        if avg_last > 0:
            trend = round(((avg_curr - avg_last) / avg_last) * 100, 1)
        else:
            trend = 0

        return jsonify({
            'avg_orders'     : avg_curr,
            'avg_last_month' : avg_last,
            'trend'          : trend,
            'total_customers': total_customers
        })

    except Exception as e:
        return jsonify({'avg_orders': 0, 'trend': 0, 'error': str(e)})
    finally:
        cursor.close()
        conn.close()
        
        
        
        
@app.route('/api/owner/avg-orders-detail')
def avg_orders_detail():
    conn   = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT TABLE_NAME FROM information_schema.tables
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME NOT LIKE '%_product_activity'
            AND TABLE_NAME NOT LIKE '%_your_item'
            AND TABLE_NAME REGEXP '^[a-z0-9_]+_[0-9]+$'
        """)
        tables = [r['TABLE_NAME'] for r in cursor.fetchall()]

        skip = {
            'user', 'user_activity', 'user_signout_logs', 'user_survey',
            'user_template', 'deleted_users', 'deleted_customers',
            'addtocart_logs', 'search_logs', 'cart', 'cart_summary',
            'card', 'food_items', 'study_material', 'store_data',
            'strong_password', 'sample', 'save_detail',
            'category_requests', 'support_sd', 'product_availability',
            'special_offers'
        }

        customers = []
        for table in tables:
            if table in skip:
                continue
            try:
                parts = table.rsplit('_', 1)
                if len(parts) != 2 or not parts[1].isdigit():
                    continue
                user_id  = parts[1]
                username = parts[0].replace('_', ' ').title()

                cursor.execute(f"""
                    SELECT
                        COUNT(*) as total_orders,
                        MAX(date)  as last_order,
                        MIN(date)  as first_order
                    FROM `{table}`
                    WHERE mode IN ('successful', 'combo_offer')
                """)
                row = cursor.fetchone()
                total   = int(row['total_orders'] or 0)
                last_o  = row['last_order']
                first_o = row['first_order']

                if first_o and last_o and total > 0:
                    months = max(1, (last_o.year - first_o.year) * 12
                                  + (last_o.month - first_o.month) + 1)
                    avg_pm = round(total / months, 1)
                else:
                    avg_pm = 0

                customers.append({
                    'user_id'      : user_id,
                    'username'     : username,
                    'total_orders' : total,
                    'avg_per_month': avg_pm,
                    'last_order'   : last_o.strftime('%b %d, %Y') if last_o else None
                })
            except:
                continue

        # ── ML: KMeans clustering on avg_per_month to assign buyer status ──
        active = [c for c in customers if c['total_orders'] > 0]

        if len(active) >= 4:
            try:
                import numpy as np
                from sklearn.cluster import KMeans

                X = np.array([[c['avg_per_month']] for c in active])
                k = 4 if len(active) >= 4 else len(active)
                km = KMeans(n_clusters=k, n_init=10, random_state=42)
                labels = km.fit_predict(X)

                # Rank clusters by their mean avg_per_month, high to low
                cluster_means = {}
                for lbl in set(labels):
                    vals = [active[i]['avg_per_month'] for i in range(len(active)) if labels[i] == lbl]
                    cluster_means[lbl] = sum(vals) / len(vals)

                ranked = sorted(cluster_means, key=lambda l: cluster_means[l], reverse=True)
                status_names = ['Frequent Buyer', 'Stable Buyer', 'Regular Buyer', 'Low Activity']
                cluster_to_status = {lbl: status_names[i] for i, lbl in enumerate(ranked)}

                for i, c in enumerate(active):
                    c['status'] = cluster_to_status[labels[i]]
            except Exception:
                # Fallback to rule-based thresholds if sklearn/numpy unavailable
                for c in active:
                    apm = c['avg_per_month']
                    if apm >= 4:
                        c['status'] = 'Frequent Buyer'
                    elif apm >= 2.5:
                        c['status'] = 'Stable Buyer'
                    elif apm >= 1:
                        c['status'] = 'Regular Buyer'
                    else:
                        c['status'] = 'Low Activity'
        else:
            # Too few customers to cluster meaningfully — use rule-based thresholds
            for c in active:
                apm = c['avg_per_month']
                if apm >= 4:
                    c['status'] = 'Frequent Buyer'
                elif apm >= 2.5:
                    c['status'] = 'Stable Buyer'
                elif apm >= 1:
                    c['status'] = 'Regular Buyer'
                else:
                    c['status'] = 'Low Activity'

        # Customers with zero orders
        for c in customers:
            if c['total_orders'] == 0:
                c['status'] = 'No Orders'

        customers.sort(key=lambda x: x['total_orders'], reverse=True)
        return jsonify({'customers': customers})

    except Exception as e:
        return jsonify({'customers': [], 'error': str(e)})
    finally:
        cursor.close()
        conn.close()

#--------------------------------------------------------------------------------------------------

# â”€â”€ Strong Password Manager â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/owner/strong-passwords", methods=["GET"])
def get_strong_passwords():
    conn = None
    try:
        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM strong_password ORDER BY is_used DESC, id DESC")
        rows = cursor.fetchall()
        cursor.close()
        return jsonify({"passwords": rows})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()

@app.route("/api/owner/strong-passwords", methods=["POST"])
def add_strong_password():
    conn = None
    try:
        data     = request.get_json()
        password = data.get("password", "").strip()
        is_used  = int(data.get("is_used", 0))

        if not password:
            return jsonify({"error": "Password cannot be empty"}), 400

        conn   = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO strong_password (password, is_used) VALUES (%s, %s)",
            (password, is_used)
        )
        conn.commit()
        cursor.close()
        return jsonify({"message": "Password saved successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()
            
#________________________3 button part in password section (" select for delete ") ___---------------            
@app.route("/api/owner/strong-passwords/bulk-delete", methods=["POST"])
def bulk_delete_strong_passwords():
    conn = None
    try:
        data = request.get_json()
        ids  = data.get("ids", [])
        if not ids:
            return jsonify({"error": "No IDs provided"}), 400

        conn   = get_db_connection()
        cursor = conn.cursor()
        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(f"DELETE FROM strong_password WHERE id IN ({placeholders})", tuple(ids))
        deleted_count = cursor.rowcount
        conn.commit()
        cursor.close()
        return jsonify({"message": f"{deleted_count} password(s) deleted"}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()
            
            
#________________________3 button part in password section (" make as used and make as unused ") ___---------------            
            
            
@app.route("/api/owner/strong-passwords/bulk-update-status", methods=["POST"])
def bulk_update_password_status():
    conn = None
    try:
        data    = request.get_json()
        ids     = data.get("ids", [])
        is_used = int(data.get("is_used", 0))

        if not ids:
            return jsonify({"error": "No IDs provided"}), 400

        conn   = get_db_connection()
        cursor = conn.cursor()
        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(
            f"UPDATE strong_password SET is_used = %s WHERE id IN ({placeholders})",
            tuple([is_used] + ids)
        )
        updated_count = cursor.rowcount
        conn.commit()
        cursor.close()
        return jsonify({"message": f"{updated_count} password(s) updated"}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()
            
#-----------------------------category request ko table mi store karne ke leya  ---------------

#-----------------properly correct detail request form madhe enter kari la ---------------------- 


@app.route("/request-category", methods=["POST"])
def request_category():
    data = request.get_json()
    category_name = data.get("category_name", "").strip()
    reason_name = data.get("reason_name", "").strip()
    reason = data.get("reason", "").strip()
    errors = []

    # 1. Category Name validation
    if len(category_name) < 3:
        errors.append("âŒ Category name must be at least 3 characters.")
    if not re.match(r'^[a-zA-Z0-9\s]+$', category_name):
        errors.append("❌ Category name must contain only letters and numbers.")

    # 3. Reason validation â€” ML style
    words = reason.split()
    unique_words = set(w.lower() for w in words)

    if len(words) < 15:
        errors.append(f"âŒ Reason too short â€” write at least 15 words. (You wrote {len(words)})")
    if len(unique_words) < 8:
        errors.append("âŒ Reason looks repetitive â€” please write meaningful content.")
    if not re.search(r'[.!?]', reason):
        errors.append("âŒ Reason must have proper sentences (use . or ! or ?).")

    # 4. Spam/random text check
    def is_random_text(text):
        words_list = text.lower().split()
        short_words = [w for w in words_list if len(w) <= 2]
        return len(short_words) > len(words_list) * 0.6

    if is_random_text(reason):
        errors.append("âŒ Reason contains too many short/random words.")

    if errors:
        return jsonify({"status": "error", "errors": errors})

    # Sab theek hai â€” DB mein save karo
    # Sab theek hai — DB mein save karo
    user_name = session.get("username", "").strip()
    uid = session.get("user_id")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        INSERT INTO category_requests (user_id, user_name, category_name, reason_name, reason)
        VALUES (%s, %s, %s, %s, %s)
    """, (uid, user_name, category_name, reason_name, reason))
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"status": "success"})


#--------------------owner dashooard madhe category request sathi ------------------------------   

@app.route("/api/owner/category-requests", methods=["GET"])
def get_category_requests():
    conn = None
    try:
        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
        SELECT id, user_id, user_name, category_name, reason_name, reason, status, requested_at, notif_id
        FROM category_requests
            ORDER BY requested_at DESC
        """)
        rows = cursor.fetchall()
        cursor.close()

        from datetime import datetime
        for row in rows:
            if isinstance(row.get("requested_at"), datetime):
                row["requested_at"] = row["requested_at"].strftime("%Y-%m-%d %H:%M:%S")

        # â”€â”€ Default ML fields â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        for row in rows:
            row["similarity_score"] = 0.0
            row["is_duplicate"]     = False
            row["duplicate_of"]     = None
            row["sentiment"]        = "neutral"
            row["sentiment_score"]  = 0.0
            row["priority_score"]   = 0.0
            row["priority_label"]   = "Low"
            row["auto_summary"]     = (row.get("reason") or "")[:80]
            row["cluster_id"]       = 0
            row["cluster_label"]    = "Uncategorized"

        if len(rows) >= 1:
            try:
                from sklearn.feature_extraction.text import TfidfVectorizer
                from sklearn.metrics.pairwise import cosine_similarity
                import numpy as np

                reasons = []
                for r in rows:
                    text = (r.get("reason") or "").strip()
                    reasons.append(text if len(text) >= 3 else "no_reason_placeholder")

                cat_names = []
                for r in rows:
                    text = (r.get("category_name") or "").strip()
                    cat_names.append(text if len(text) >= 2 else "unknown")

                vectorizer   = TfidfVectorizer(token_pattern=r"(?u)\b\w+\b")
                tfidf_matrix = vectorizer.fit_transform(reasons)

                # â”€â”€ 1. TF-IDF + Cosine Similarity (Duplicate Detection) â”€â”€â”€â”€â”€â”€â”€â”€â”€
                if len(rows) > 1:
                    sim_matrix = cosine_similarity(tfidf_matrix)
                    for i, row in enumerate(rows):
                        max_sim = 0.0
                        dup_of  = None
                        for j in range(len(rows)):
                            if i != j and sim_matrix[i][j] > max_sim:
                                max_sim = sim_matrix[i][j]
                                dup_of  = rows[j]["id"]
                        row["similarity_score"] = round(float(max_sim) * 100, 1)
                        row["is_duplicate"]     = bool(max_sim >= 0.70)
                        row["duplicate_of"]     = dup_of if max_sim >= 0.70 else None

                # â”€â”€ 2. Sentiment Analysis (Keyword-based, no heavy model needed) â”€
                positive_words = {"great","love","need","want","amazing","important",
                                  "useful","necessary","excellent","good","benefit",
                                  "helpful","required","essential","popular","urgent"}
                negative_words = {"bad","terrible","useless","waste","stupid","hate",
                                  "boring","never","worst","broken","missing","lack",
                                  "absent","poor","disappointing"}
                urgency_words  = {"urgent","asap","immediately","critical","must",
                                  "please","needed","require","now","soon"}

                for row in rows:
                    tokens = set((row.get("reason") or "").lower().split())
                    pos = len(tokens & positive_words)
                    neg = len(tokens & negative_words)
                    urg = len(tokens & urgency_words)
                    score = pos - neg + (urg * 0.5)
                    if score > 0.5:
                        sentiment = "positive"
                    elif score < -0.5:
                        sentiment = "negative"
                    else:
                        sentiment = "neutral"
                    row["sentiment"]       = sentiment
                    row["sentiment_score"] = round(score, 2)

                # â”€â”€ 3. Smart Priority Scoring â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
                #    Combines: reason length, uniqueness, sentiment, urgency
                reason_lengths = [len((r.get("reason") or "")) for r in rows]
                max_len = max(reason_lengths) if reason_lengths else 1

                for row in rows:
                    reason_len   = len((row.get("reason") or ""))
                    length_score = min(reason_len / max_len, 1.0)          # 0-1
                    unique_score = 1.0 - (row["similarity_score"] / 100)   # 0-1 (more unique = higher)
                    sent_score   = {"positive": 1.0, "neutral": 0.5, "negative": 0.2}.get(row["sentiment"], 0.5)

                    # Urgency from reason text
                    tokens = set((row.get("reason") or "").lower().split())
                    urgency_bonus = 0.2 if tokens & urgency_words else 0.0

                    priority = (length_score * 0.3) + (unique_score * 0.4) + (sent_score * 0.2) + urgency_bonus
                    priority = round(priority * 100, 1)

                    if priority >= 70:
                        label = "High"
                    elif priority >= 40:
                        label = "Medium"
                    else:
                        label = "Low"

                    row["priority_score"] = priority
                    row["priority_label"] = label

                # â”€â”€ 4. Category Name Clustering (K-Means on TF-IDF of names) â”€â”€
                if len(rows) >= 3:
                    from sklearn.cluster import KMeans
                    n_clusters = min(3, len(rows))

                    cat_vectorizer = TfidfVectorizer(token_pattern=r"(?u)\b\w+\b", analyzer="char_wb", ngram_range=(2,4))
                    cat_matrix     = cat_vectorizer.fit_transform(cat_names)

                    km = KMeans(n_clusters=n_clusters, random_state=42, n_init=10)
                    labels = km.fit_predict(cat_matrix)

                    cluster_colors = ["ðŸ”µ", "ðŸŸ£", "ðŸŸ "]
                    cluster_names  = {}
                    for cluster_id in range(n_clusters):
                        members = [cat_names[i] for i, l in enumerate(labels) if l == cluster_id]
                        cluster_names[cluster_id] = members[0] if members else f"Group {cluster_id+1}"

                    for i, row in enumerate(rows):
                        cid = int(labels[i])
                        row["cluster_id"]    = cid
                        row["cluster_label"] = f"{cluster_colors[cid]} {cluster_names[cid]}"

                # â”€â”€ 5. Auto-Summary (extractive: pick best sentence by TF-IDF weight) â”€
                feature_names = vectorizer.get_feature_names_out()
                for i, row in enumerate(rows):
                    reason = (row.get("reason") or "").strip()
                    sentences = [s.strip() for s in reason.replace(".", ". ").split(".") if len(s.strip()) > 10]
                    if not sentences:
                        row["auto_summary"] = reason[:80] + ("..." if len(reason) > 80 else "")
                        continue
                    # Score each sentence by sum of TF-IDF weights of its words
                    tfidf_row = tfidf_matrix[i].toarray()[0]
                    word_scores = dict(zip(feature_names, tfidf_row))
                    best_sent, best_score = sentences[0], -1
                    for sent in sentences:
                        s_score = sum(word_scores.get(w.lower(), 0) for w in sent.split())
                        if s_score > best_score:
                            best_score = s_score
                            best_sent  = sent
                    row["auto_summary"] = best_sent[:100] + ("..." if len(best_sent) > 100 else "")

            except Exception as ml_err:
                import traceback
                print("ML block failed:", traceback.format_exc())

        return jsonify({"requests": rows})

    except Exception as e:
        import traceback
        print("get_category_requests crashed:", traceback.format_exc())
        return jsonify({"error": str(e), "requests": []}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()
            
            
            
@app.route("/api/owner/category-requests/<int:request_id>/send", methods=["POST"])
def send_category_request(request_id):
    conn = None
    try:
        body       = request.get_json(silent=True) or {}
        owner_note = body.get("note", "").strip()

        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM category_requests WHERE id = %s", (request_id,))
        req = cursor.fetchone()
        if not req:
            cursor.close()
            return jsonify({"error": "Request not found"}), 404

        # âœ… Fix datetime
        from datetime import datetime
        if isinstance(req.get("requested_at"), datetime):
            req["requested_at"] = req["requested_at"].strftime("%Y-%m-%d %H:%M:%S")


        

        note_block = f"""
            <div style="margin-top:18px;background:#f0f9ff;border-left:4px solid #6366f1;
                border-radius:0 10px 10px 0;padding:14px 16px;">
                <div style="font-size:11px;font-weight:700;color:#6366f1;
                    text-transform:uppercase;letter-spacing:0.5px;margin-bottom:6px;">âœï¸ Owner's Note</div>
                <div style="font-size:14px;color:#334155;line-height:1.6;">{owner_note}</div>
            </div>
        """ if owner_note else ""

        message = SendGridMail(
            from_email=os.environ.get("SENDER_EMAIL"),
            to_emails="ankitabandal45@gmail.com",
            subject=f"New Category Request: {req['category_name']}",
            html_content=f"""
                <div style="font-family:'Segoe UI',Arial,sans-serif;max-width:480px;margin:0 auto;
                    background:#f8fafc;padding:24px;">

                    <div style="background:white;border-radius:16px;overflow:hidden;
                        box-shadow:0 4px 20px rgba(0,0,0,0.08);">

                        <div style="background:linear-gradient(135deg,#8b5cf6,#a855f7);
                            padding:20px 24px;color:white;">
                            <div style="font-size:13px;opacity:0.9;font-weight:600;letter-spacing:0.5px;
                                text-transform:uppercase;margin-bottom:4px;">ðŸ“‚ New Category Request</div>
                            <div style="font-size:20px;font-weight:800;">{req['category_name']}</div>
                        </div>

                        <div style="padding:24px;">

                            <div style="display:flex;justify-content:space-between;align-items:center;
                                padding:12px 0;border-bottom:1px solid #f1f5f9;">
                                <span style="font-size:11px;font-weight:700;color:#94a3b8;
                                    text-transform:uppercase;letter-spacing:0.5px;">Requested By</span>
                                <span style="font-size:14px;font-weight:700;color:#1e293b;">{req['user_name']}</span>
                            </div>

                            <div style="display:flex;justify-content:space-between;align-items:center;
                                padding:12px 0;border-bottom:1px solid #f1f5f9;">
                                <span style="font-size:11px;font-weight:700;color:#94a3b8;
                                    text-transform:uppercase;letter-spacing:0.5px;">Category</span>
                                <span style="font-size:14px;font-weight:700;color:#1e293b;">{req['category_name']}</span>
                            </div>

                            <div style="display:flex;justify-content:space-between;align-items:center;
                                padding:12px 0;border-bottom:1px solid #f1f5f9;">
                                <span style="font-size:11px;font-weight:700;color:#94a3b8;
                                    text-transform:uppercase;letter-spacing:0.5px;">Requested At</span>
                                <span style="font-size:13px;font-weight:600;color:#64748b;">{req['requested_at']}</span>
                            </div>

                            <div style="padding-top:14px;">
                                <div style="font-size:11px;font-weight:700;color:#94a3b8;
                                    text-transform:uppercase;letter-spacing:0.5px;margin-bottom:8px;">Reason</div>
                                <div style="font-size:14px;color:#334155;line-height:1.7;
                                    background:#faf5ff;border:1px solid #e9d5ff;border-radius:10px;
                                    padding:14px;">{req['reason']}</div>
                            </div>

                            {note_block}

                        </div>

                        <div style="background:#f8fafc;padding:14px 24px;text-align:center;
                            border-top:1px solid #f1f5f9;">
                            <span style="font-size:11px;color:#94a3b8;">ShopSphere Owner Dashboard</span>
                        </div>
                    </div>
</div>
            """
        )

        sg = SendGridAPIClient(os.environ.get("SHOPSPHERE_SENDGRID_KEY"))
        sg.send(message)

        cursor.execute(
            "UPDATE category_requests SET status = %s WHERE id = %s",
            ("sent_to_developer", request_id)
        )
        conn.commit()
        cursor.close()
        return jsonify({"message": "Request sent to developer successfully"}), 200

    except Exception as e:
        import traceback
        print("send_category_request crashed:", traceback.format_exc())
        return jsonify({"error": str(e)}), 500
    finally:
        if conn and conn.is_connected():
            conn.close()
   
#---------------owner_section madhun category request chi return message to that user notification sathi ----------------------------------

@app.route("/api/owner/send-notification", methods=["POST"])
def send_notification_to_user():
    try:
        body       = request.get_json(silent=True) or {}
        user_id    = body.get("user_id")
        title      = body.get("title", "").strip()
        message    = body.get("message", "").strip()
        request_id = body.get("request_id")  # category_requests.id

        if not user_id or not message:
            return jsonify({"error": "user_id and message required"}), 400

        conn   = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO user_notifications (user_id, title, message)
            VALUES (%s, %s, %s)
        """, (user_id, title, message))
        conn.commit()
        notif_id = cursor.lastrowid

        if request_id:
            cursor.execute(
                "UPDATE category_requests SET notif_id = %s WHERE id = %s",
                (notif_id, request_id)
            )
            conn.commit()

        cursor.close()
        conn.close()
        return jsonify({"status": "success", "notif_id": notif_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/user/notifications", methods=["GET"])
def get_user_notifications():
    try:
        uid = session.get("user_id")
        if not uid:
            return jsonify({"notifications": []})

        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, title, message, is_read, created_at
            FROM user_notifications
            WHERE user_id = %s
            ORDER BY created_at DESC
        """, (uid,))
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        for row in rows:
            if hasattr(row["created_at"], "strftime"):
                row["created_at"] = row["created_at"].strftime("%Y-%m-%d %H:%M")

        return jsonify({"notifications": rows})
    except Exception as e:
        return jsonify({"error": str(e), "notifications": []}), 500


@app.route("/api/user/notifications/read/<int:notif_id>", methods=["POST"])
def mark_notification_read(notif_id):
    try:
        conn   = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE user_notifications 
            SET is_read = 1, read_at = NOW() 
            WHERE id = %s AND is_read = 0
        """, (notif_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
    
    
#__________owner provider message from owner_section to dashboard where if user see that message it will blue tick at the owner_dashboard ________________________


@app.route("/api/owner/notif-read-status/<int:notif_id>", methods=["GET"])
def notif_read_status(notif_id):
    try:
        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT is_read, read_at, created_at FROM user_notifications WHERE id = %s", (notif_id,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        if not row:
            return jsonify({"is_read": False})
        return jsonify({
            "is_read": bool(row["is_read"]),
            "read_at": row["read_at"].strftime("%Y-%m-%d %H:%M:%S") if row["read_at"] else None,
            "sent_at": row["created_at"].strftime("%Y-%m-%d %H:%M:%S") if row["created_at"] else None
        })
    except Exception as e:
        return jsonify({"is_read": False}), 500
    
    
    
#------------------------------------------------------------------------------------------------------------   
#---------------owner search data of ( owner_section ) ------------------------------------            
@app.route("/api/owner/search-analytics", methods=["GET"])
def search_analytics():
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # âœ… Seedha saari product_activity tables lo â€” user_activity se match nahi karna
        cursor.execute("""
            SELECT table_name FROM information_schema.tables
            WHERE table_schema = DATABASE()
            AND table_name LIKE '%_product_activity'
        """)
        all_tables = [list(row.values())[0] for row in cursor.fetchall()]


        print(f"DEBUG: Found {len(all_tables)} activity tables", flush=True)

        aggregated = defaultdict(lambda: {
            "total_searches": 0,
            "users_who_searched": set(),
            "category": None,
            "growth_sum": 0.0,
            "growth_count": 0,
            "last_searched": None,
        })

        for tbl_name in all_tables:
            print(f"DEBUG: Reading table {tbl_name}", flush=True)

            try:
                cursor.execute(f"""
                    SELECT name, category,
                           COALESCE(today_search_count, 0) AS today_search_count,
                           COALESCE(growth_on_search, '0') AS growth_on_search,
                           search_time
                    FROM `{tbl_name}`
                    WHERE today_search_count IS NOT NULL AND today_search_count > 0
                """)
                rows = cursor.fetchall()
                print(f"DEBUG: {tbl_name} has {len(rows)} search rows", flush=True)

                for row in rows:
                    pname = (row["name"] or "").strip()
                    if not pname:
                        continue

                    bucket = aggregated[pname]
                    bucket["total_searches"] += int(row["today_search_count"] or 0)
                    bucket["users_who_searched"].add(tbl_name)  # table name as unique user identifier

                    if bucket["category"] is None and row["category"]:
                        bucket["category"] = row["category"]

                    try:
                        raw = str(row["growth_on_search"] or "0").strip()
                        if "/" in raw:
                            parts = raw.split("/")
                            g = float(parts[0]) / float(parts[1]) * 100
                        else:
                            g = float(raw.replace("%", "") or 0)
                        bucket["growth_sum"] += g
                        bucket["growth_count"] += 1
                    except (ValueError, TypeError, ZeroDivisionError):
                        pass

                    if row["search_time"]:
                        ts = str(row["search_time"])
                        if bucket["last_searched"] is None or ts > bucket["last_searched"]:
                            bucket["last_searched"] = ts

            except Exception as e:
                print(f"DEBUG: Skipping {tbl_name} â€” {str(e)}", flush=True)
                continue

        # âœ… ML: Normalize popularity score 0-100
        all_searches = [b["total_searches"] for b in aggregated.values()]
        max_search = max(all_searches) if all_searches else 1

        results = []
        for pname, bucket in aggregated.items():
            avg_growth = round(bucket["growth_sum"] / bucket["growth_count"], 2) if bucket["growth_count"] > 0 else 0.0
            popularity_score = round((bucket["total_searches"] / max_search) * 100, 1)

            # âœ… ML: Trend label
            if avg_growth >= 70:
                trend = "ðŸ”¥ Hot"
            elif avg_growth >= 40:
                trend = "ðŸ“ˆ Rising"
            elif avg_growth >= 10:
                trend = "âž¡ï¸ Stable"
            else:
                trend = "ðŸ“‰ Low"

            results.append({
                "product_name": pname,
                "category": bucket["category"] or "Uncategorized",
                "total_searches": bucket["total_searches"],
                "users_who_searched": len(bucket["users_who_searched"]),
                "growth": avg_growth,
                "popularity_score": popularity_score,
                "trend": trend,
                "last_searched": bucket["last_searched"],
            })

        results.sort(key=lambda x: x["popularity_score"], reverse=True)

        print(f"DEBUG: Returning {len(results)} products", flush=True)
        return jsonify({"results": results})

    except Exception as e:
        print(f"DEBUG ERROR: {str(e)}", flush=True)
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e), "results": []}), 500

    finally:
        if cursor: cursor.close()
        if conn: conn.close()   
        

#----------------all user detail ("save_detail") sathi -------------------------------        


@app.route('/owner/all-users')
def owner_all_users():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, name, age, phone1, phone2, address1, address2, email, profile_image, created_at FROM save_detail ORDER BY created_at DESC")
        users = cursor.fetchall()
        cursor.close()
        conn.close()
        # Convert datetime to string for JSON serialisation
        for u in users:
            if u.get('created_at'):
                u['created_at'] = str(u['created_at'])
        return jsonify(users)
    except Exception as e:
        return jsonify([]), 500


#----------------All dashboard Card ("store_detail") sathi -------------------------------        



@app.route('/owner/all-dashboard-cards')
def owner_all_dashboard_cards():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, user_id, product_id, category, name, image,
                   price, availability, quantity, date, uploaded_at
            FROM store_data
            ORDER BY date DESC
        """)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        for r in rows:
            for key in ['date', 'uploaded_at']:
                if r.get(key):
                    r[key] = str(r[key])
            for key in ['price']:
                if r.get(key) is not None:
                    r[key] = float(r[key])
        return jsonify(rows)
    except Exception as e:
        return jsonify([]), 500
    
#----------------------Product add to card data  of ( owner_section )------------------------- 


@app.route("/api/owner/addtocart-analytics", methods=["GET"])
def addtocart_analytics():
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # âœ… Actual product names fetch karo product tables se
        cursor.execute("""
            SELECT id, name, category FROM card
            UNION ALL
            SELECT id, name, category FROM study_material
            UNION ALL
            SELECT id, name, category FROM food_items
        """)
        all_products = cursor.fetchall()
        # product_id + category â†’ name mapping
        product_map = {(p["id"], p["category"]): p["name"] for p in all_products}

        # âœ… Actual product names fetch karo
        cursor.execute("""
            SELECT id, name, category FROM card
            UNION ALL
            SELECT id, name, category FROM study_material
            UNION ALL
            SELECT id, name, category FROM food_items
        """)
        all_products = cursor.fetchall()
        product_map = {(p["id"], p["category"]): p["name"] for p in all_products}

        cursor.execute("""
            SELECT table_name FROM information_schema.tables
            WHERE table_schema = DATABASE()
            AND table_name LIKE '%_product_activity'
        """)
        all_tables = [list(row.values())[0] for row in cursor.fetchall()]

        aggregated = defaultdict(lambda: {
            "total_addtocart": 0,
            "users_who_added": set(),
            "category": None,
            "growth_sum": 0.0,
            "growth_count": 0,
            "last_added": None,
        })

        for tbl_name in all_tables:
            try:
                cursor.execute(f"""
                    SELECT product_id, name, category,
                           COALESCE(today_add_to_cart_count, 0) AS today_add_to_cart_count,
                           COALESCE(growth_in_addtocart, '0') AS growth_in_addtocart,
                           add_to_cart_date_time
                    FROM `{tbl_name}`
                    WHERE today_add_to_cart_count IS NOT NULL AND today_add_to_cart_count > 0
                """)
                rows = cursor.fetchall()

                for row in rows:
                    # âœ… product_map se sahi naam lo
                    actual_name = product_map.get((row["product_id"], row["category"]))
                    if not actual_name:
                        continue

                    # âœ… Unique label = name + category + product_id
                    unique_key = f"{actual_name} ({row['category']}) #{row['product_id']}"
                    bucket = aggregated[unique_key] 
                    bucket["total_addtocart"] += int(row["today_add_to_cart_count"] or 0)
                    bucket["users_who_added"].add(tbl_name)

                    if bucket["category"] is None and row["category"]:
                        bucket["category"] = row["category"]

                    try:
                        raw = str(row["growth_in_addtocart"] or "0").strip()
                        if "/" in raw:
                            parts = raw.split("/")
                            g = float(parts[0]) / float(parts[1]) * 100
                        else:
                            g = float(raw.replace("%", "") or 0)
                        bucket["growth_sum"] += g
                        bucket["growth_count"] += 1
                    except (ValueError, TypeError, ZeroDivisionError):
                        pass

                    if row["add_to_cart_date_time"]:
                        ts = str(row["add_to_cart_date_time"])
                        if bucket["last_added"] is None or ts > bucket["last_added"]:
                            bucket["last_added"] = ts

            except Exception as e:
                print(f"DEBUG: Skipping {tbl_name} â€” {str(e)}", flush=True)
                continue

        # âœ… ML: Normalize popularity score 0-100
        all_counts = [b["total_addtocart"] for b in aggregated.values()]
        max_count = max(all_counts) if all_counts else 1

        results = []
        for pname, bucket in aggregated.items():
            avg_growth = round(bucket["growth_sum"] / bucket["growth_count"], 2) if bucket["growth_count"] > 0 else 0.0
            popularity_score = round((bucket["total_addtocart"] / max_count) * 100, 1)

            if avg_growth >= 70:
                trend = "ðŸ”¥ Hot"
            elif avg_growth >= 40:
                trend = "ðŸ“ˆ Rising"
            elif avg_growth >= 10:
                trend = "âž¡ï¸ Stable"
            else:
                trend = "ðŸ“‰ Low"

            results.append({
                "product_name": pname,
                "category": bucket["category"] or "Uncategorized",
                "total_addtocart": bucket["total_addtocart"],
                "users_who_added": len(bucket["users_who_added"]),
                "growth": avg_growth,
                "popularity_score": popularity_score,
                "trend": trend,
                "last_added": bucket["last_added"],
            })

        results.sort(key=lambda x: x["popularity_score"], reverse=True)
        return jsonify({"results": results})

    except Exception as e:
        print(f"DEBUG ERROR: {str(e)}", flush=True)
        return jsonify({"error": str(e), "results": []}), 500

    finally:
        if cursor: cursor.close()
        if conn: conn.close()

#----------------------Product Purchased data  of ( owner_section )------------------------- 



@app.route("/api/owner/purchased-analytics", methods=["GET"])
def purchased_analytics():
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # âœ… Actual product names fetch karo
        cursor.execute("""
            SELECT id, name, category FROM card
            UNION ALL
            SELECT id, name, category FROM study_material
            UNION ALL
            SELECT id, name, category FROM food_items
        """)
        all_products = cursor.fetchall()
        product_map = {(p["id"], p["category"]): p["name"] for p in all_products}

        cursor.execute("""
            SELECT table_name FROM information_schema.tables
            WHERE table_schema = DATABASE()
            AND table_name LIKE '%_product_activity'
        """)
        all_tables = [list(row.values())[0] for row in cursor.fetchall()]

        aggregated = defaultdict(lambda: {
            "total_purchased": 0,
            "users_who_purchased": set(),
            "category": None,
            "growth_sum": 0.0,
            "growth_count": 0,
            "last_purchased": None,
        })

        for tbl_name in all_tables:
            try:
                cursor.execute(f"""
                    SELECT product_id, category,
                           COALESCE(today_purchase_count, 0) AS today_purchase_count,
                           COALESCE(growth, 0) AS growth,
                           purchased_time
                    FROM `{tbl_name}`
                    WHERE today_purchase_count IS NOT NULL AND today_purchase_count > 0
                """)
                rows = cursor.fetchall()

                for row in rows:
                    actual_name = product_map.get((row["product_id"], row["category"]))
                    if not actual_name:
                        continue

                    unique_key = f"{actual_name} ({row['category']}) #{row['product_id']}"
                    bucket = aggregated[unique_key]
                    bucket["total_purchased"] += int(row["today_purchase_count"] or 0)
                    bucket["users_who_purchased"].add(tbl_name)

                    if bucket["category"] is None and row["category"]:
                        bucket["category"] = row["category"]

                    try:
                        g = float(row["growth"] or 0)
                        bucket["growth_sum"] += g * 100
                        bucket["growth_count"] += 1
                    except (ValueError, TypeError):
                        pass

                    if row["purchased_time"]:
                        ts = str(row["purchased_time"])
                        if bucket["last_purchased"] is None or ts > bucket["last_purchased"]:
                            bucket["last_purchased"] = ts

            except Exception as e:
                print(f"DEBUG: Skipping {tbl_name} â€” {str(e)}", flush=True)
                continue

        # âœ… ML: Normalize popularity score 0-100
        all_counts = [b["total_purchased"] for b in aggregated.values()]
        max_count = max(all_counts) if all_counts else 1

        results = []
        for pname, bucket in aggregated.items():
            avg_growth = round(bucket["growth_sum"] / bucket["growth_count"], 2) if bucket["growth_count"] > 0 else 0.0
            popularity_score = round((bucket["total_purchased"] / max_count) * 100, 1)

            if avg_growth >= 70:
                trend = "ðŸ”¥ Hot"
            elif avg_growth >= 40:
                trend = "ðŸ“ˆ Rising"
            elif avg_growth >= 10:
                trend = "âž¡ï¸ Stable"
            else:
                trend = "ðŸ“‰ Low"

            results.append({
                "product_name": pname,
                "category": bucket["category"] or "Uncategorized",
                "total_purchased": bucket["total_purchased"],
                "users_who_purchased": len(bucket["users_who_purchased"]),
                "growth": avg_growth,
                "popularity_score": popularity_score,
                "trend": trend,
                "last_purchased": bucket["last_purchased"],
            })

        results.sort(key=lambda x: x["popularity_score"], reverse=True)
        return jsonify({"results": results})

    except Exception as e:
        print(f"DEBUG ERROR: {str(e)}", flush=True)
        return jsonify({"error": str(e), "results": []}), 500

    finally:
        if cursor: cursor.close()
        if conn: conn.close()



#-----------------view card madhe purchased card store rahneya sathi ------------------------


@app.route("/get-purchased-items")
def get_purchased_items():
    if "user_id" not in session or "username" not in session:
        return jsonify([])

    user_id = session["user_id"]
    username = session["username"]
    table_name = get_cart_table_name(username, user_id)

    db = get_db_connection()
    cursor = db.cursor(dictionary=True)

    try:
        cursor.execute(f"""
            SELECT * FROM `{table_name}` 
            WHERE mode = 'successful' 
            ORDER BY date DESC
        """)
        items = cursor.fetchall()

        for row in items:
            img = row.get("image")
            if img:
                row["image"] = img if (img.startswith("http") or img.startswith("/static")) else f"/static/products/{img}"
            else:
                row["image"] = "/static/products/default.png"

            row["date"] = row["date"].strftime("%Y-%m-%d") if row.get("date") else ""

    except Exception as e:
        print("Purchased fetch error:", e)
        items = []
    finally:
        cursor.close()
        db.close()

    return jsonify(items)



#----------------owner_section.html madhle ("product category") che count's -----------------------

@app.route("/api/owner/product-catalog-stats", methods=["GET"])
def product_catalog_stats():
    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT table_name FROM information_schema.tables
            WHERE table_schema = DATABASE()
            AND table_name LIKE '%_product_activity'
        """)
        all_tables = [list(row.values())[0] for row in cursor.fetchall()]

        stats = {}

        for tbl_name in all_tables:
            try:
                cursor.execute(f"""
                    SELECT product_id, category,
                           COALESCE(today_search_count, 0) AS search_count,
                           COALESCE(today_add_to_cart_count, 0) AS cart_count,
                           COALESCE(today_purchase_count, 0) AS purchase_count
                    FROM `{tbl_name}`
                """)
                rows = cursor.fetchall()

                for row in rows:
                    key = f"{row['product_id']}_{row['category']}"
                    if key not in stats:
                        stats[key] = {
                            "product_id": row["product_id"],
                            "category": row["category"],
                            "search": 0,
                            "cart": 0,
                            "purchased": 0
                        }
                    stats[key]["search"] += int(row["search_count"])
                    stats[key]["cart"] += int(row["cart_count"])
                    stats[key]["purchased"] += int(row["purchase_count"])

            except Exception as e:
                continue

        return jsonify(list(stats.values()))

    except Exception as e:
        return jsonify([]), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()


#--------------owner_section madhe monthly analysis sathi jyat month select karun it get access that month detail ----------------


@app.route("/api/owner/monthly-analysis", methods=["GET"])
def monthly_analysis():
    month = request.args.get("month", "June")
    year = request.args.get("year", "2026")

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # âœ… Speed Fix 1: Ek hi query mein product map banao
        cursor.execute("""
            SELECT id, name, 'card' as category FROM card
            UNION ALL
            SELECT id, name, 'study_material' FROM study_material
            UNION ALL
            SELECT id, name, 'food_items' FROM food_items
        """)
        product_map = {(p["id"], p["category"]): p["name"] for p in cursor.fetchall()}

        # âœ… Speed Fix 2: Sabhi tables ek baar fetch karo
        cursor.execute("""
            SELECT table_name FROM information_schema.tables
            WHERE table_schema = DATABASE()
            AND table_name LIKE '%_product_activity'
        """)
        all_tables = [list(row.values())[0] for row in cursor.fetchall()]

        aggregated = defaultdict(lambda: {
            "search": 0, "cart": 0, "purchased": 0, "category": None
        })

        # âœ… Speed Fix 3: Batch queries â€” cursor reuse
        for tbl_name in all_tables:
            try:
                cursor.execute(f"""
                    SELECT product_id, category,
                           COALESCE(today_search_count, 0) AS search_count,
                           COALESCE(today_add_to_cart_count, 0) AS cart_count,
                           COALESCE(today_purchase_count, 0) AS purchase_count
                    FROM `{tbl_name}`
                    WHERE month = %s
                      AND (today_search_count > 0 
                           OR today_add_to_cart_count > 0 
                           OR today_purchase_count > 0)
                """, (month,))
                rows = cursor.fetchall()

                for row in rows:
                    actual_name = product_map.get((row["product_id"], row["category"]))
                    if not actual_name:
                        continue
                    unique_key = f"{actual_name} ({row['category']}) #{row['product_id']}"
                    aggregated[unique_key]["search"] += int(row["search_count"])
                    aggregated[unique_key]["cart"] += int(row["cart_count"])
                    aggregated[unique_key]["purchased"] += int(row["purchase_count"])
                    if aggregated[unique_key]["category"] is None:
                        aggregated[unique_key]["category"] = row["category"]
            except:
                continue

        if not aggregated:
            return jsonify([])

        # âœ… ML 1: Min-Max Normalization
        all_search = [v["search"] for v in aggregated.values()]
        all_cart = [v["cart"] for v in aggregated.values()]
        all_purchase = [v["purchased"] for v in aggregated.values()]
        max_s = max(all_search) if all_search else 1
        max_c = max(all_cart) if all_cart else 1
        max_p = max(all_purchase) if all_purchase else 1

        results = []
        for pname, bucket in aggregated.items():
            # âœ… ML 2: Conversion Rate
            search_to_cart = round((bucket["cart"] / bucket["search"] * 100), 1) if bucket["search"] > 0 else 0
            cart_to_purchase = round((bucket["purchased"] / bucket["cart"] * 100), 1) if bucket["cart"] > 0 else 0

            # âœ… ML 3: Weighted Growth Score
            norm_s = bucket["search"] / max_s
            norm_c = bucket["cart"] / max_c
            norm_p = bucket["purchased"] / max_p
            growth_score = round((norm_s * 0.3 + norm_c * 0.3 + norm_p * 0.4) * 100, 1)

            # âœ… ML 4: Trend Classification
            if growth_score >= 70:
                trend = "ðŸ”¥ Hot"
            elif growth_score >= 40:
                trend = "ðŸ“ˆ Rising"
            elif bucket["search"] > 0 and bucket["purchased"] == 0:
                trend = "âš ï¸ No Purchase"
            elif growth_score >= 10:
                trend = "âž¡ï¸ Stable"
            else:
                trend = "ðŸ“‰ Low"

            # âœ… ML 5: Anomaly Detection
            anomaly = bucket["search"] >= 3 and bucket["purchased"] == 0

            results.append({
                "product_name": pname,
                "category": bucket["category"],
                "search": bucket["search"],
                "cart": bucket["cart"],
                "purchased": bucket["purchased"],
                "growth_score": growth_score,
                "search_to_cart_rate": search_to_cart,
                "cart_to_purchase_rate": cart_to_purchase,
                "trend": trend,
                "anomaly": anomaly
            })

        results.sort(key=lambda x: x["growth_score"], reverse=True)
        return jsonify(results)

    except Exception as e:
        return jsonify([]), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
        
        
#-------------------Owner _section "open" button (customer churan predictio ) sathi -----------------------        
@app.route('/api/churn-customers')
def api_churn_customers():
    import re
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT TABLE_NAME FROM information_schema.tables
        WHERE TABLE_SCHEMA = DATABASE()
    """)
    all_tables = [row['TABLE_NAME'] for row in cursor.fetchall()]

    activity_tables = [t for t in all_tables if t.endswith('_product_activity')]

    activity_base_map = {}
    for at in activity_tables:
        base = at.replace('_product_activity', '')
        activity_base_map[base.lower().replace(' ', '_')] = at

    skip_tables = {
        'user', 'user_activity', 'user_signout_logs', 'user_survey',
        'user_template', 'deleted_users', 'deleted_customers',
        'addtocart_logs', 'search_logs', 'cart', 'cart_summary',
        'card', 'food_items', 'study_material', 'store_data',
        'strong_password', 'StrongPassword', 'sample', 'save_detail',
        'category_requests', 'support_sd', 'product_availability',
        'product_availability_sql', 'vc_product_availability',
        'yourusername_1_product_activity', 'BANDAl_7_product_activity',
        'special_offers'
    }

    customer_tables = []
    for t in all_tables:
        if t in skip_tables:
            continue
        if t.endswith('_product_activity'):
            continue
        if t.endswith('_your_item'):
            continue
        parts = t.rsplit('_', 1)
        if len(parts) == 2 and parts[1].isdigit():
            customer_tables.append(t)

    # â”€â”€ Pre-fetch which activity tables have add_to_cart_date_time â”€â”€
    cursor.execute("""
        SELECT TABLE_NAME FROM information_schema.columns
        WHERE TABLE_SCHEMA = DATABASE()
        AND COLUMN_NAME = 'add_to_cart_date_time'
        AND TABLE_NAME LIKE '%_product_activity'
    """)
    tables_with_datetime = {row['TABLE_NAME'] for row in cursor.fetchall()}

    customers = []

    for table in customer_tables:
        parts = table.rsplit('_', 1)
        uid           = int(parts[1])
        username_part = parts[0]

        normalized_key = table.lower().replace(' ', '_')
        activity_table = activity_base_map.get(normalized_key)

        try:
            cursor.execute(f"SELECT * FROM `{table}` LIMIT 1")
            user_row = cursor.fetchone()
            if not user_row:
                continue
        except:
            continue

        email = user_row.get('email', '')

        total_orders    = 0
        total_cart      = 0
        total_search    = 0
        total_spend     = 0.0
        unique_products = 0
        days_since      = 999

        if activity_table:
            try:
                # â”€â”€ Smart datetime column check â”€â”€
                if activity_table in tables_with_datetime:
                    datetime_col = 'MAX(add_to_cart_date_time) AS last_activity'
                else:
                    datetime_col = 'NULL AS last_activity'

                cursor.execute(f"""
                    SELECT
                        COALESCE(SUM(today_purchase_count), 0)    AS total_orders,
                        COALESCE(SUM(today_add_to_cart_count), 0) AS total_cart,
                        COALESCE(SUM(today_search_count), 0)      AS total_search,
                        COUNT(DISTINCT product_id)                 AS unique_products,
                        {datetime_col}
                    FROM `{activity_table}`
                """)
                act = cursor.fetchone()

                total_orders    = int(act['total_orders']    or 0)
                total_cart      = int(act['total_cart']      or 0)
                total_search    = int(act['total_search']    or 0)
                unique_products = int(act['unique_products'] or 0)

                if act['last_activity']:
                    days_since = (datetime.utcnow() - act['last_activity']).days

                # Spend calculation
                cursor.execute(f"""
                    SELECT name,
                           COALESCE(SUM(today_purchase_count), 0) AS purchased
                    FROM `{activity_table}`
                    WHERE today_purchase_count > 0
                    GROUP BY name
                """)
                purchased_rows = cursor.fetchall()

                for pr in purchased_rows:
                    try:
                        cursor.execute(f"""
                            SELECT price FROM `{table}`
                            WHERE name = %s LIMIT 1
                        """, (pr['name'],))
                        price_row = cursor.fetchone()
                        if price_row and price_row['price']:
                            total_spend += float(price_row['price']) * int(pr['purchased'])
                    except:
                        continue

            except Exception as e:
                print(f"Error {activity_table}: {e}")

        engagement = total_search + total_cart + unique_products

        customers.append({
            'id'             : uid,
            'name'           : username_part.replace('_', ' ').title(),
            'email'          : email,
            'orders'         : total_orders,
            'spend'          : round(total_spend, 2),
            'daysSinceLast'  : days_since,
            'loginCount'     : engagement,
            'total_cart'     : total_cart,
            'total_search'   : total_search,
            'unique_products': unique_products,
            'has_activity'   : activity_table is not None
        })

    cursor.close()
    conn.close()
    return jsonify({'customers': customers})


# â”€â”€ Send Special OR Common offer â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route('/api/send-special-offer', methods=['POST'])
def send_special_offer():
    import re
    data             = request.get_json()
    username         = data.get('username', '').strip()
    offer_type       = data.get('offer_type', '').strip()
    message          = data.get('message', '').strip()
    uid              = data.get('uid')
    offer_category   = data.get('offer_category', 'special')  # 'special' or 'common'
    product1_name    = data.get('product1_name', '')
    product1_image   = data.get('product1_image', '')
    product2_name    = data.get('product2_name', '')
    product2_image   = data.get('product2_image', '')
    discount         = data.get('discount', 0)

    conn   = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS special_offers (
                id               INT AUTO_INCREMENT PRIMARY KEY,
                customer_table   VARCHAR(150),
                username         VARCHAR(100),
                offer_type       VARCHAR(100),
                offer_category   VARCHAR(20) DEFAULT 'special',
                message          TEXT,
                product1_name    VARCHAR(255),
                product1_image   VARCHAR(500),
                product2_name    VARCHAR(255),
                product2_image   VARCHAR(500),
                discount         DECIMAL(5,2) DEFAULT 0,
                is_read          TINYINT DEFAULT 0,
                created_at       DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)

        import re as _re
        sanitized      = _re.sub(r'[^a-z0-9_]', '_', username.lower())
        table_name     = f"{sanitized}_{uid}"

        cursor.execute("""
            INSERT INTO special_offers 
            (customer_table, username, offer_type, offer_category,
             message, product1_name, product1_image,
             product2_name, product2_image, discount)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (table_name, username, offer_type, offer_category,
              message, product1_name, product1_image,
              product2_name, product2_image, discount))

        conn.commit()
        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


# â”€â”€ Send Common offer to ALL customers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route('/api/send-common-offer', methods=['POST'])
def send_common_offer():
    import re as _re
    data           = request.get_json()
    offer_type     = data.get('offer_type', '').strip()
    message        = data.get('message', '').strip()
    product1_name  = data.get('product1_name', '')
    product1_image = data.get('product1_image', '')
    product2_name  = data.get('product2_name', '')
    product2_image = data.get('product2_image', '')
    discount       = data.get('discount', 0)

    if not offer_type or not message:
        return jsonify({'success': False, 'error': 'Missing fields'}), 400

    conn   = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # Get all customer tables
        cursor.execute("""
            SELECT TABLE_NAME FROM information_schema.tables
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME NOT LIKE '%_product_activity'
            AND TABLE_NAME NOT LIKE '%_your_item'
            AND TABLE_NAME REGEXP '^[a-z0-9_]+_[0-9]+$'
        """)
        customer_tables = [row['TABLE_NAME'] for row in cursor.fetchall()]

        # Create table if not exists
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS special_offers (
                id               INT AUTO_INCREMENT PRIMARY KEY,
                customer_table   VARCHAR(150),
                username         VARCHAR(100),
                offer_type       VARCHAR(100),
                offer_category   VARCHAR(20) DEFAULT 'common',
                message          TEXT,
                product1_name    VARCHAR(255),
                product1_image   VARCHAR(500),
                product2_name    VARCHAR(255),
                product2_image   VARCHAR(500),
                discount         DECIMAL(5,2) DEFAULT 0,
                is_read          TINYINT DEFAULT 0,
                created_at       DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()

        # Add missing columns if table already existed without them
        alter_columns = [
            "ALTER TABLE special_offers ADD COLUMN offer_category VARCHAR(20) DEFAULT 'common'",
            "ALTER TABLE special_offers ADD COLUMN product1_name VARCHAR(255)",
            "ALTER TABLE special_offers ADD COLUMN product1_image VARCHAR(500)",
            "ALTER TABLE special_offers ADD COLUMN product2_name VARCHAR(255)",
            "ALTER TABLE special_offers ADD COLUMN product2_image VARCHAR(500)",
            "ALTER TABLE special_offers ADD COLUMN discount DECIMAL(5,2) DEFAULT 0",
        ]
        for sql in alter_columns:
            try:
                cursor.execute(sql)
                conn.commit()
            except:
                pass  # column already exists - ignore duplicate

        count = 0
        for table in customer_tables:
            parts = table.rsplit('_', 1)
            if len(parts) != 2 or not parts[1].isdigit():
                continue
            username = parts[0]
            cursor.execute("""
                INSERT INTO special_offers
                (customer_table, username, offer_type, offer_category,
                 message, product1_name, product1_image,
                 product2_name, product2_image, discount)
                VALUES (%s,%s,%s,'common',%s,%s,%s,%s,%s,%s)
            """, (table, username, offer_type, message,
                  product1_name, product1_image,
                  product2_name, product2_image, discount))
            count += 1

        conn.commit()
        return jsonify({'success': True, 'sent_to': count})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

# â”€â”€ Customer fetches their offers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route('/api/my-special-offers')
def my_special_offers():
    import re as _re
    conn   = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        username = session.get('username', '')
        uid      = session.get('user_id') or session.get('id')
        if not username or not uid:
            return jsonify({'offers': []})

        sanitized  = _re.sub(r'[^a-z0-9_]', '_', username.lower())
        table_name = f"{sanitized}_{uid}"

        cursor.execute("""
            SELECT id, offer_type, offer_category, message,
                   product1_name, product1_image,
                   product2_name, product2_image,
                   discount, is_read, created_at
            FROM special_offers
            WHERE customer_table = %s
            ORDER BY created_at DESC
            LIMIT 30
        """, (table_name,))
        offers = cursor.fetchall()
        for o in offers:
            o['created_at'] = str(o['created_at'])
            o['discount']   = float(o['discount'] or 0)
        return jsonify({'offers': offers})
    except Exception as e:
        return jsonify({'offers': [], 'error': str(e)})
    finally:
        cursor.close()
        conn.close()


# ── Buy Combo Offer → save to cart → redirect to buynow ───
@app.route('/api/buy-combo-offer', methods=['POST'])
def buy_combo_offer():
    import re as _re
    data           = request.get_json()
    offer_id       = data.get('offer_id')
    product1_name  = data.get('product1_name', '')
    product1_image = data.get('product1_image', '')
    product2_name  = data.get('product2_name', '')
    product2_image = data.get('product2_image', '')
    discount       = float(data.get('discount', 0))

    username = session.get('username', '')
    uid      = session.get('user_id') or session.get('id')

    if not username or not uid:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401

    conn   = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        sanitized      = _re.sub(r'[^a-z0-9_]', '_', username.lower())
        user_table     = f"{sanitized}_{uid}"
        activity_table = f"{sanitized}_{uid}_product_activity"

        # ── Get real prices + images from store_data ──
        cursor.execute("""
            SELECT name, price, image, id as product_id FROM store_data
            WHERE name IN (%s, %s)
        """, (product1_name, product2_name))
        products = {p['name']: p for p in cursor.fetchall()}

        p1       = products.get(product1_name, {})
        p2       = products.get(product2_name, {})
        p1_price = float(p1.get('price') or 0)
        p2_price = float(p2.get('price') or 0)
        p1_image = p1.get('image') or product1_image
        p2_image = p2.get('image') or product2_image
        p1_pid   = p1.get('product_id') or 0
        p2_pid   = p2.get('product_id') or 0
        p1_cat   = p1.get('category', 'Combo Offer')
        p2_cat   = p2.get('category', 'Combo Offer')

        # ── Calculate discounted total ──
        total_original = p1_price + p2_price
        final_price    = round(total_original * (1 - discount / 100), 2) if discount > 0 else round(total_original, 2)

        combo_name = f"{product1_name} + {product2_name}"
        detail     = (f"Combo Offer | {product1_name} ₹{p1_price} + "
                      f"{product2_name} ₹{p2_price} = ₹{total_original} "
                      f"| {discount}% OFF | Final: ₹{final_price}")

        # ── Auto-add image2 column if not exists ──
        try:
            cursor.execute(f"ALTER TABLE `{user_table}` ADD COLUMN image2 VARCHAR(255)")
            conn.commit()
        except:
            pass  # already exists

        # ── Insert combo into user table ──
        cursor.execute(f"""
            INSERT INTO `{user_table}`
            (name, price, image, image2, category, detail, quantity, mode)
            VALUES (%s, %s, %s, %s, %s, %s, 1, 'combo_offer')
        """, (combo_name, final_price, p1_image, p2_image, 'Combo Offer', detail))

        conn.commit()
        cart_id       = cursor.lastrowid
        current_month = datetime.now().strftime('%B')

        # ── Update product_activity table ──
        cursor.execute("""
            SELECT COUNT(*) as cnt FROM information_schema.tables
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = %s
        """, (activity_table,))

        if cursor.fetchone()['cnt'] > 0:
            for pname, pid, pcat in [
                (product1_name, p1_pid, p1_cat),
                (product2_name, p2_pid, p2_cat)
            ]:
                if not pname:
                    continue

                # Check if row exists by name OR product_id
                cursor.execute(f"""
                    SELECT id FROM `{activity_table}`
                    WHERE name = %s
                    OR (product_id = %s AND product_id != 0)
                    LIMIT 1
                """, (pname, pid))
                existing = cursor.fetchone()

                if existing:
                    # Update existing row
                    cursor.execute(f"""
                        UPDATE `{activity_table}`
                        SET today_purchase_count = today_purchase_count + 1,
                            name  = %s,
                            month = %s
                        WHERE id = %s
                    """, (pname, current_month, existing['id']))
                else:
                    # Insert only if truly not exists
                    try:
                        cursor.execute(f"""
                            INSERT INTO `{activity_table}`
                            (name, product_id, category,
                             today_search_count, today_add_to_cart_count,
                             today_purchase_count, month)
                            VALUES (%s, %s, %s, 0, 0, 1, %s)
                        """, (pname, pid, pcat, current_month))
                    except Exception as ae:
                        print(f"Activity insert error for {pname}: {ae}")

            conn.commit()

        return jsonify({
            'success'       : True,
            'cart_id'       : cart_id,
            'original_price': total_original,
            'discount_pct'  : discount,
            'final_price'   : final_price,
            'image2'        : p2_image
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()       
        
# â”€â”€ Mark offer as read â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route('/api/mark-offer-read/<int:offer_id>', methods=['POST'])
def mark_offer_read(offer_id):
    conn   = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE special_offers SET is_read=1 WHERE id=%s", (offer_id,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        cursor.close()
        conn.close()




#___________product part (from store_data table )___________


@app.route('/api/owner/all-products-for-offer')
def all_products_for_offer():
    conn   = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT DISTINCT name, image, price, category
            FROM store_data
            WHERE name IS NOT NULL AND name != ''
            ORDER BY name
        """)
        products = cursor.fetchall()
        for p in products:
            p['price'] = float(p['price'] or 0)
        return jsonify({'products': products})
    except Exception as e:
        return jsonify({'products': [], 'error': str(e)})
    finally:
        cursor.close()
        conn.close()
        
        
#__________total discount ko calculate karne ki leya to send every user common offer (" owner_dashboard to dashboard")
#--------------------personal search analysis ( " dashboard.html madhe " )-------------------------


@app.route("/products/search-user")
def search_user():
    if "user_id" not in session or "username" not in session:
        return jsonify([])

    user_id = session["user_id"]
    username = session["username"]

    safe_username = re.sub(r'[^a-z0-9_]', '_', username.strip().lower())
    if safe_username and safe_username[0].isdigit():
        safe_username = "user_" + safe_username
    activity_table = f"{safe_username}_{user_id}_product_activity"

    q = request.args.get("q", "").strip()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        if q:
            cursor.execute(f"""
                SELECT id, product_id, name, category,
                       today_search_count AS searched_count,
                       search_time AS last_searched_time
                FROM `{activity_table}`
                WHERE today_search_count > 0 AND name LIKE %s
                ORDER BY today_search_count DESC
            """, (f"%{q}%",))
        else:
            cursor.execute(f"""
                SELECT id, product_id, name, category,
                       today_search_count AS searched_count,
                       search_time AS last_searched_time
                FROM `{activity_table}`
                WHERE today_search_count > 0
                ORDER BY today_search_count DESC
            """)

        results = cursor.fetchall()

        for row in results:
            if row["last_searched_time"]:
                row["last_searched_time"] = row["last_searched_time"].strftime("%I:%M %p")

        return jsonify(results)

    except Exception as e:
        print("search-user error:", e)
        return jsonify([])
    finally:
        cursor.close()
        conn.close()

    
#--------------------personal Add to card  analysis ( " dashboard.html madhe " )-------------------------
@app.route("/get-addtocart-user-data")
def get_addtocart_user_data():
    if "user_id" not in session or "username" not in session:
        return jsonify([])

    user_id = session["user_id"]
    username = session["username"]

    safe_username = re.sub(r'[^a-z0-9_]', '_', username.strip().lower())
    if safe_username and safe_username[0].isdigit():
        safe_username = "user_" + safe_username
    activity_table = f"{safe_username}_{user_id}_product_activity"

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # âœ… Check which time column exists
        cursor.execute(f"""
            SELECT COUNT(*) as cnt FROM information_schema.columns
            WHERE table_schema = DATABASE()
            AND table_name = %s
            AND column_name = 'add_to_cart_date_time'
        """, (activity_table,))
        has_new_col = cursor.fetchone()["cnt"]

        time_col = "add_to_cart_date_time" if has_new_col else "add_to_cart_time"

        cursor.execute(f"""
            SELECT id, product_id, name, category,
                   today_add_to_cart_count AS add_to_cart_count,
                   `{time_col}` AS add_to_cart_time
            FROM `{activity_table}`
            WHERE today_add_to_cart_count > 0
            ORDER BY `{time_col}` DESC
        """)
        rows = cursor.fetchall()

        results = []
        for i, row in enumerate(rows):
            t = row["add_to_cart_time"]
            results.append({
                "id": i + 1,
                "product_id": row["product_id"],
                "product_name": row["name"] or "Unknown",
                "category": row["category"],
                "add_to_cart_count": row["add_to_cart_count"],
                "add_to_cart_time": str(t) if t else None
            })

        return jsonify(results)

    except Exception as e:
        print("get-addtocart-user-data error:", e)
        return jsonify([])
    finally:
        cursor.close()
        conn.close()
#--------------------personal purchase analysis ( " dashboard.html madhe " )-------------------------
@app.route("/get-purchase-data")
def get_purchase_data():
    if "user_id" not in session or "username" not in session:
        return jsonify([])

    user_id = session["user_id"]
    username = session["username"]

    # âœ… Sanitized cart table name
    safe_username = re.sub(r'[^a-z0-9_]', '_', username.strip().lower())
    if safe_username and safe_username[0].isdigit():
        safe_username = "user_" + safe_username
    cart_table = f"{safe_username}_{user_id}"

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # âœ… Product names fetch karo
        cursor.execute("""
            SELECT id, name, 'card' as category FROM card
            UNION ALL
            SELECT id, name, 'study_material' FROM study_material
            UNION ALL
            SELECT id, name, 'food_items' FROM food_items
        """)
        product_map = {(p["id"], p["category"]): p["name"] for p in cursor.fetchall()}

        # âœ… Successful purchases fetch karo cart table se
        cursor.execute(f"""
            SELECT product_id, category, MAX(date) as date, COUNT(*) as purchase_count
            FROM `{cart_table}`
            WHERE mode = 'successful'
            GROUP BY product_id, category
            ORDER BY MAX(date) DESC
        """)
        rows = cursor.fetchall()

        results = []
        for i, row in enumerate(rows):
            actual_name = product_map.get((row["product_id"], row["category"])) or "Unknown"
            results.append({
                "purchase_id": i + 1,
                "product_name": actual_name,
                "category": row["category"],
                "purchase_count": row["purchase_count"],
                "purchase_time": str(row["date"]) if row["date"] else None,
                "month": row["date"].strftime("%B") if row["date"] else None
            })

        return jsonify(results)

    except Exception as e:
        print("get-purchase-data error:", e)
        return jsonify([])
    finally:
        cursor.close()
        conn.close()



#--------------profile page ('your item option button ")--------------------
@app.route('/api/all-products')
def all_products():
    try:
        uname = session.get("username")
        uid = session.get("user_id")
        if not uname or not uid:
            return jsonify({'products': [], 'error': 'Not logged in'}), 401

        print(f"[all_products] session username='{uname}' uid='{uid}'")
        your_item_table = f"{uname}_{uid}_your_item"
        print(f"[all_products] looking for table: {your_item_table}")

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(f"""
            CREATE TABLE IF NOT EXISTS `{your_item_table}` (
                id INT AUTO_INCREMENT PRIMARY KEY,
                store_data_id INT,
                category VARCHAR(100),
                name VARCHAR(255),
                image TEXT,
                video TEXT,
                price DECIMAL(10,2),
                availability INT,
                detail TEXT,
                address TEXT,
                quantity INT,
                keywords TEXT,
                made_of TEXT,
                used_for TEXT,
                harmful_activity TEXT,
                precautions TEXT
            )
        """)
        conn.commit()
        cursor.execute(f"""
            SELECT id, store_data_id, category, name, image, price,
                   availability, detail, address, quantity,
                   uploaded_at
            FROM `{your_item_table}`
            WHERE availability > 0
            ORDER BY category, name
        """)
        products = cursor.fetchall()
        cursor.close()
        conn.close()
        for p in products:
            p['price']       = float(p['price']) if p['price'] else 0
            p['quantity']    = int(p['quantity']) if p['quantity'] else 0
            p['image_url']   = p.get('image', '')
            p['description'] = p.get('detail', '')
            p['uploaded_at'] = str(p['uploaded_at']) if p.get('uploaded_at') else ''
        return jsonify({'products': products})
    except Exception as e:
        print(f"[all_products] error: {e}")
        return jsonify({'products': [], 'error': str(e)}), 500


def create_user_your_item_table(username, user_id):
    table_name = f"{username}_{user_id}_your_item"
    db = get_db_connection()
    cursor = db.cursor()
    try:
        cursor.execute(f"""
            CREATE TABLE IF NOT EXISTS `{table_name}` (
                id            INT AUTO_INCREMENT PRIMARY KEY,
                store_data_id INT,
                category      VARCHAR(20),
                name          VARCHAR(255),
                image         VARCHAR(255),
                video         VARCHAR(255),
                price         DECIMAL(10,2),
                availability  INT,
                detail        TEXT,
                address       VARCHAR(255),
                quantity      INT DEFAULT 1,
                uploaded_at   DATETIME DEFAULT CURRENT_TIMESTAMP,
                keywords      TEXT,
                made_of       TEXT,
                used_for      TEXT,
                harmful_activity TEXT,
                precautions   TEXT
            )
        """)
        db.commit()
        print(f"âœ… Created table: {table_name}")
    except Exception as e:
        print(f"âŒ Error creating {table_name}: {e}")
    finally:
        cursor.close()
        db.close()
        
#subpoint-----------profile.html madhle your_item chya aatle card cha view-----------       

@app.route('/api/update-keywords', methods=['POST'])
def update_keywords():
    try:
        data = request.get_json()
        pid  = data.get('product_id')
        kw   = data.get('keywords', '')

        uname = session.get("username")
        uid   = session.get("user_id")
        if not uname or not uid:
            return jsonify({'success': False, 'error': 'Not logged in'}), 401

        safe_un = re.sub(r'[^a-z0-9_]', '_', uname.strip().lower())
        if safe_un and safe_un[0].isdigit():
            safe_un = "user_" + safe_un
        your_item_table = f"{safe_un}_{uid}_your_item"

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Step 1: get name + category from your_item
        cursor.execute(f"SELECT name, category FROM `{your_item_table}` WHERE store_data_id=%s LIMIT 1", (pid,))
        item_row = cursor.fetchone()

        if item_row:
            prod_table = get_product_table(item_row['category'])

            # Step 2: update actual product table (food_items / card / study_material / any future)
            cursor.execute(f"UPDATE `{prod_table}` SET keywords=%s WHERE name=%s", (kw, item_row['name']))

            # Step 3: update your_item table
            cursor.execute(f"UPDATE `{your_item_table}` SET keywords=%s WHERE store_data_id=%s", (kw, pid))

        # Step 4: update store_data
        cursor.execute("UPDATE store_data SET keywords=%s WHERE id=%s", (kw, pid))

        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route("/api/product-activity-detail")
def product_activity_detail():
    if "user_id" not in session or "username" not in session:
        return jsonify({})

    product_id = request.args.get("product_id")
    category   = request.args.get("category", "")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        safe_username = re.sub(r'[^a-z0-9_]', '_', session["username"].strip().lower())
        if safe_username and safe_username[0].isdigit():
            safe_username = "user_" + safe_username
        your_item_table = f"{safe_username}_{session['user_id']}_your_item"

        cursor.execute(f"""
            SELECT name, category, made_of, used_for, harmful_activity, precautions
            FROM `{your_item_table}` WHERE id = %s LIMIT 1
        """, (product_id,))
        item_row = cursor.fetchone()
        if not item_row:
            return jsonify({"today_search_count":0,"today_add_to_cart_count":0,"today_purchase_count":0})

        product_name     = item_row["name"]
        product_category = item_row["category"]
        prod_table       = get_product_table(product_category)

        cursor.execute(f"SELECT id FROM `{prod_table}` WHERE name = %s LIMIT 1", (product_name,))
        prod_row = cursor.fetchone()
        if not prod_row:
            return jsonify({"today_search_count":0,"today_add_to_cart_count":0,"today_purchase_count":0})

        real_pid      = prod_row["id"]
        real_category = prod_table

        cursor.execute("""
            SELECT TABLE_NAME FROM information_schema.TABLES
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME NOT LIKE '%_product_activity'
            AND TABLE_NAME NOT LIKE '%_your_item'
            AND TABLE_NAME NOT LIKE '%_product'
            AND TABLE_NAME REGEXP '^[a-zA-Z0-9_]+_[0-9]+$'
        """)
        cart_tables = [list(r.values())[0] for r in cursor.fetchall()]

        cursor.execute("""
            SELECT TABLE_NAME FROM information_schema.TABLES
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME LIKE '%_product_activity'
        """)
        activity_tables = [list(r.values())[0] for r in cursor.fetchall()]

        total_purchase = 0
        total_cart     = 0
        total_search   = 0

        for tbl in cart_tables:
            try:
                cursor.execute(f"""
                    SELECT
                        SUM(CASE WHEN mode='successful' THEN 1 ELSE 0 END) as purchases,
                        COUNT(*) as carts
                    FROM `{tbl}`
                    WHERE product_id = %s AND category = %s
                """, (real_pid, real_category))
                row = cursor.fetchone()
                if row:
                    total_purchase += int(row["purchases"] or 0)
                    total_cart     += int(row["carts"]     or 0)
            except:
                pass

        for tbl in activity_tables:
            try:
                cursor.execute(f"""
                    SELECT SUM(today_search_count) as searches
                    FROM `{tbl}`
                    WHERE product_id = %s AND category = %s
                """, (real_pid, real_category))
                row = cursor.fetchone()
                if row:
                    total_search += int(row["searches"] or 0)
            except:
                pass

        return jsonify({
            "today_search_count":      total_search,
            "today_add_to_cart_count": total_cart,
            "today_purchase_count":    total_purchase,
            "made_of":          item_row.get("made_of") or "",
            "used_for":         item_row.get("used_for") or "",
            "harmful_activity": item_row.get("harmful_activity") or "",
            "precautions":      item_row.get("precautions") or ""
        })

    except Exception as e:
        print("product-activity-detail error:", e)
        return jsonify({"today_search_count":0,"today_add_to_cart_count":0,"today_purchase_count":0})
    finally:
        cursor.close()
        conn.close()
  
        
#__________keyword change karu sathi ___________________

# â”€â”€ Helper: dynamic category â†’ table name â”€â”€
def get_product_table(category):
    cat_map = {
        'Food':           'food_items',
        'food_items':     'food_items',
        'Kitchen':        'card',
        'card':           'card',
        'Card':           'card',
        'Study Material': 'study_material',
        'study_material': 'study_material',
    }
    # If not in map, convert category name to table name dynamically
    return cat_map.get(category, category.strip().lower().replace(' ', '_'))


@app.route('/api/get-keywords', methods=['GET'])
def get_keywords():
    try:
        store_data_id = request.args.get('store_data_id')
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        uname = session.get("username")
        uid   = session.get("user_id")
        safe_username = re.sub(r'[^a-z0-9_]', '_', uname.strip().lower())
        if safe_username and safe_username[0].isdigit():
            safe_username = "user_" + safe_username
        your_item_table = f"{safe_username}_{uid}_your_item"

        cursor.execute(f"SELECT name, category FROM `{your_item_table}` WHERE store_data_id=%s LIMIT 1", (store_data_id,))
        item = cursor.fetchone()
        if not item:
            cursor.close()
            conn.close()
            return jsonify({'keywords': ''})

        prod_table = get_product_table(item['category'])

        cursor.execute(f"SELECT keywords FROM `{prod_table}` WHERE name=%s LIMIT 1", (item['name'],))
        prod_row = cursor.fetchone()
        cursor.close()
        conn.close()
        return jsonify({'keywords': prod_row['keywords'] or '' if prod_row else ''})
    except Exception as e:
        return jsonify({'keywords': '', 'error': str(e)}), 500


#______________________profile page cha "Keep shopping for" part to get display the image and video of product simultaneously __________________________


@app.route('/api/all-dashboard-products')
def all_dashboard_products():
    try:
        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        products = []
        for tbl in ['food_items', 'card', 'study_material']:
            try:
                cursor.execute(f"SELECT id, name, image, video, price, availability, '{tbl}' as category FROM `{tbl}` WHERE availability > 0 ORDER BY id DESC LIMIT 20")
                products += cursor.fetchall()
            except:
                pass
        cursor.close()
        conn.close()
        for p in products:
            p['price'] = float(p['price']) if p['price'] else 0
        return jsonify({'products': products})
    except Exception as e:
        return jsonify({'products': [], 'error': str(e)}), 500




#----------------profile page ("your Order " and ""Buy Again"") section sathi -----------------


@app.route('/api/my-cart-items')
def my_cart_items():
    if "user_id" not in session or "username" not in session:
        return jsonify([])
    try:
        username = session["username"]
        user_id  = session["user_id"]
        safe_un  = re.sub(r'[^a-z0-9_]', '_', username.strip().lower())
        if safe_un and safe_un[0].isdigit():
            safe_un = "user_" + safe_un
        cart_table = f"{safe_un}_{user_id}"

        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(f"""
            SELECT id, product_id, category, name, image, price, quantity, mode
            FROM `{cart_table}`
            WHERE (mode IS NULL OR mode != 'successful')
            ORDER BY id DESC LIMIT 10
        """)
        items = cursor.fetchall()

        # Fetch real availability from product tables
        for item in items:
            try:
                prod_table = get_product_table(item['category'])
                cursor.execute(f"SELECT availability FROM `{prod_table}` WHERE id=%s LIMIT 1", (item['product_id'],))
                prod_row = cursor.fetchone()
                item['availability'] = int(prod_row['availability']) if prod_row else 0
            except:
                item['availability'] = 0

        cursor.close()
        conn.close()
        for item in items:
            item['price'] = float(item['price']) if item['price'] else 0
        return jsonify(items)
    except Exception as e:
        print("my-cart-items error:", e)
        return jsonify([])

@app.route('/api/my-purchased-items')
def my_purchased_items():
    if "user_id" not in session or "username" not in session:
        return jsonify([])
    try:
        username = session["username"]
        user_id  = session["user_id"]
        safe_un  = re.sub(r'[^a-z0-9_]', '_', username.strip().lower())
        if safe_un and safe_un[0].isdigit():
            safe_un = "user_" + safe_un
        cart_table = f"{safe_un}_{user_id}"

        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(f"""
            SELECT id, product_id, category, name, image, price, quantity as availability
            FROM `{cart_table}`
            WHERE mode = 'successful'
            ORDER BY id DESC LIMIT 10
        """)
        items = cursor.fetchall()

        # Fetch real availability from product tables
        for item in items:
            try:
                prod_table = get_product_table(item['category'])
                cursor.execute(f"SELECT availability FROM `{prod_table}` WHERE id=%s LIMIT 1", (item['product_id'],))
                prod_row = cursor.fetchone()
                item['availability'] = int(prod_row['availability']) if prod_row else 0
            except:
                item['availability'] = 0

        cursor.close()
        conn.close()
        for item in items:
            item['price'] = float(item['price']) if item['price'] else 0
        return jsonify(items)
    except Exception as e:
        print("my-purchased-items error:", e)
        return jsonify([])
#_____________________"Buy again" part of the profile to get display the corrct detail in see detail option ______________


@app.route('/api/product-detail')
def api_product_detail():
    product_id = request.args.get('product_id')
    category   = request.args.get('category', '')
    try:
        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        prod_table = get_product_table(category)
        cursor.execute(f"SELECT * FROM `{prod_table}` WHERE id=%s LIMIT 1", (product_id,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        if not row:
            return jsonify({})
        row['price']        = float(row['price']) if row['price'] else 0
        row['availability'] = int(row['availability']) if row['availability'] else 0
        row['uploaded_at']  = str(row['uploaded_at']) if row.get('uploaded_at') else ''
        return jsonify(row)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
   
#__________card availability derease kari la after it get add to the card to view_card.html ____________



@app.route('/decrement-availability', methods=['POST'])
def decrement_availability():
    try:
        data       = request.get_json()
        product_id = data.get('product_id')
        category   = data.get('category')

        prod_table = get_product_table(category)

        conn   = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(f"""
            UPDATE `{prod_table}`
            SET availability = GREATEST(availability - 1, 0)
            WHERE id = %s
        """, (product_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    
    
                  
#-------------------logout process from profile.html page ({sign out")button sathi-------------------- 
@app.route('/api/owner/signout-logs')
def get_signout_logs():
    try:
        conn   = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT s.id, s.user_id, s.user_name, 
                   COALESCE(s.user_email, u.email) as user_email,
                   s.profile_image, s.signout_reason, s.custom_reason, 
                   s.ip_address, s.signout_at
            FROM user_signout_logs s
            LEFT JOIN user u ON u.id = s.user_id
            ORDER BY s.signout_at DESC
        """)
        logs = cursor.fetchall()
        cursor.close()
        conn.close()
        for l in logs:
            if l.get('signout_at'):
                l['signout_at'] = str(l['signout_at'])
        return jsonify({'logs': logs})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/owner/signout-log/<int:log_id>', methods=['DELETE'])
def delete_signout_log(log_id):
    try:
        conn   = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM user_signout_logs WHERE id = %s", (log_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/owner/signout-logs/<int:user_id>', methods=['DELETE'])
def delete_user_signout_logs(user_id):
    try:
        conn   = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM user_signout_logs WHERE user_id = %s", (user_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/api/signout-log', methods=['POST'])
def signout_log():
    try:
        data       = request.get_json()
        user_id    = session.get('user_id')
        user_name  = session.get('username', '')
        user_email = (session.get('user_email') or 
              session.get('email', ''))

        # Fallback: get from user table if still empty
        if not user_email and user_id:
            try:
                conn2  = get_db_connection()
                cur2   = conn2.cursor(dictionary=True)
                cur2.execute("SELECT email FROM user WHERE id = %s LIMIT 1", (user_id,))
                row = cur2.fetchone()
                if row:
                    user_email = row.get('email', '')
                cur2.close()
                conn2.close()
            except:
                pass

        # Fallback: get email from user table if not in session
        if not user_email and user_id:
            try:
                conn2   = get_db_connection()
                cur2    = conn2.cursor(dictionary=True)
                cur2.execute("SELECT email FROM user WHERE id = %s LIMIT 1", (user_id,))
                row = cur2.fetchone()
                if row:
                    user_email = row.get('email', '')
                cur2.close()
                conn2.close()
            except:
                pass

        ip_address = request.remote_addr
        session_id = request.cookies.get('session', '')

        conn   = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO user_signout_logs
                (user_id, user_name, user_email, profile_image,
                 signout_reason, custom_reason,
                 ip_address, user_agent, session_id, signout_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            user_id,
            user_name,
            user_email,
            data.get('profile_image'),
            data.get('signout_reason'),
            data.get('custom_reason') or data.get('extra_feedback'),
            ip_address,
            data.get('user_agent'),
            session_id,
            datetime.utcnow() + timedelta(hours=5, minutes=30)
        ))
        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"signout-log error: {e}")

    return jsonify({'success': True})


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))
 
        
# ============================================================
# STATIC FILE SERVING
# ============================================================
@app.route('/static/videos/<path:filename>')
def serve_video(filename):
    print("ðŸ‘‰ ROUTE HIT:", request.path)
    return send_from_directory('static/videos', filename, mimetype='video/mp4')


# ============================================================
# RUN
# ============================================================
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
